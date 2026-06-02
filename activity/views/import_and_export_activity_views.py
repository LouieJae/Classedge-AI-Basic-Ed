from django.shortcuts import render, redirect
from activity.models import Activity
import csv
import io
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, permission_required
from subject.models import Subject
from django.views.decorators.http import require_POST
from django.db import transaction
from django.contrib import messages
from activity.models import Activity, ActivityType, QuizType, ActivityQuestion, QuestionChoice

@login_required
@permission_required('activity.view_activity', raise_exception=True)
def import_and_export_activity_page(request):
    from accounts.utils.pagination_utils import (
        paginate_queryset,
        search_queryset,
        get_pagination_context,
    )

    search_query = request.GET.get('search', '').strip()

    activities = Activity.objects.all().select_related('subject', 'activity_type')

    # Search
    search_fields = [
        'subject__subject_name',
        'activity_name',
        'activity_type__name',
    ]
    activities = search_queryset(activities, search_query, search_fields)

    # Pagination
    page_obj, paginator = paginate_queryset(activities, request, items_per_page=10)
    pagination_context = get_pagination_context(page_obj, request)

    context = {
        'page_obj': page_obj,
        'search_query': search_query,
    }
    context.update(pagination_context)

    return render(request, 'activity/assessments/import-and-export-assessment-page.html', context)


@login_required
@permission_required('activity.view_activity', raise_exception=True)
def export_activities(request):
    """
    Export activities with questions and choices as CSV
    Format: Flattened structure with activity + question + choice data
    """
    subject_id = request.GET.get('subject_id')
    
    # Build queryset
    activities = Activity.objects.select_related(
        'subject', 'activity_type'
    ).prefetch_related(
        'activityquestion_set__choices'
    )
    
    if subject_id:
        activities = activities.filter(subject_id=subject_id)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="activities_export.csv"'
    
    writer = csv.writer(response)
    
    # Write headers
    headers = [
        'subject_name', 'activity_name', 'activity_type',
        'max_score', 'passing_score', 'passing_score_type', 'retake_method',
        'time_duration', 'show_score', 'remedial', 'max_retake',
        'classroom_mode', 'shuffle_questions', 'activity_instruction',
        'start_time', 'end_time', 'question_order', 'question_text',
        'quiz_type', 'question_score', 'correct_answer', 'choice_order',
        'choice_text', 'is_left_side'
    ]
    writer.writerow(headers)
    
    # Write data
    for activity in activities:
        questions = activity.activityquestion_set.all()
        
        if not questions.exists():
            # Write activity row without questions
            writer.writerow([
                activity.subject.subject_name,
                activity.activity_name,
                activity.activity_type.name if activity.activity_type else '',
                activity.max_score or 100,
                activity.passing_score,
                activity.passing_score_type,
                activity.retake_method,
                activity.time_duration,
                activity.show_score,
                activity.remedial,
                activity.max_retake,
                activity.classroom_mode,
                activity.shuffle_questions,
                activity.activity_instruction or '',
                activity.start_time.strftime('%Y-%m-%d %H:%M:%S') if activity.start_time else '',
                activity.end_time.strftime('%Y-%m-%d %H:%M:%S') if activity.end_time else '',
                '', '', '', '', '', '', '', ''
            ])
        else:
            # Write activity with questions and choices
            for question_order, question in enumerate(questions, 1):
                choices = question.choices.all()
                
                if not choices.exists():
                    # Write question row without choices
                    writer.writerow([
                        activity.subject.subject_name,
                        activity.activity_name,
                        activity.activity_type.name if activity.activity_type else '',
                        activity.max_score or 100,
                        activity.passing_score,
                        activity.passing_score_type,
                        activity.retake_method,
                        activity.time_duration,
                        activity.show_score,
                        activity.remedial,
                        activity.max_retake,
                        activity.classroom_mode,
                        activity.shuffle_questions,
                        activity.activity_instruction or '',
                        activity.start_time.strftime('%Y-%m-%d %H:%M:%S') if activity.start_time else '',
                        activity.end_time.strftime('%Y-%m-%d %H:%M:%S') if activity.end_time else '',
                        question_order,
                        question.question_text,
                        question.quiz_type.name if question.quiz_type else '',
                        question.score or 0,
                        question.correct_answer,
                        '', '', ''
                    ])
                else:
                    # Write question with choices
                    for choice_order, choice in enumerate(choices, 1):
                        writer.writerow([
                            activity.subject.subject_name,
                            activity.activity_name,
                            activity.activity_type.name if activity.activity_type else '',
                            activity.max_score or 100,
                            activity.passing_score,
                            activity.passing_score_type,
                            activity.retake_method,
                            activity.time_duration,
                            activity.show_score,
                            activity.remedial,
                            activity.max_retake,
                            activity.classroom_mode,
                            activity.shuffle_questions,
                            activity.activity_instruction or '',
                            activity.start_time.strftime('%Y-%m-%d %H:%M:%S') if activity.start_time else '',
                            activity.end_time.strftime('%Y-%m-%d %H:%M:%S') if activity.end_time else '',
                            question_order,
                            question.question_text,
                            question.quiz_type.name if question.quiz_type else '',
                            question.score or 0,
                            question.correct_answer,
                            choice_order,
                            choice.choice_text,
                            choice.is_left_side
                        ])
    
    return response


@login_required
@permission_required('activity.add_activity', raise_exception=True)
@require_POST
def import_activities(request):
    """
    Import activities from CSV file
    Expected format: Same as export format (without subject_code)
    """
    import_file = request.FILES.get('import_file')
    if not import_file:
        messages.error(request, 'No file provided')
        return redirect('import_and_export_activity_page')

    try:
        # Read CSV file
        content = import_file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))

        imported_count = 0
        updated_count = 0
        errors = []

        # Group rows by activity (using subject_name + activity_name as key)
        activities_data = {}

        for row_num, row in enumerate(reader, start=2):  # Start at 2 (header is row 1)
            try:
                subject_name = row.get('subject_name', '').strip()
                activity_name = row.get('activity_name', '').strip()

                if not subject_name or not activity_name:
                    errors.append(f"Row {row_num}: Missing subject_name or activity_name")
                    continue

                # Create unique key for activity
                activity_key = f"{subject_name}|{activity_name}"

                if activity_key not in activities_data:
                    activities_data[activity_key] = {
                        'subject_name': subject_name,
                        'activity_name': activity_name,
                        'activity_data': {
                            'activity_type': row.get('activity_type', '').strip(),
                            'max_score': float(row.get('max_score', 100) or 100),
                            'passing_score': float(row.get('passing_score', 0) or 0),
                            'passing_score_type': row.get('passing_score_type', 'percentage').strip(),
                            'retake_method': row.get('retake_method', 'highest').strip(),
                            'time_duration': int(row.get('time_duration', 0) or 0),
                            'show_score': row.get('show_score', 'False').lower() == 'true',
                            'remedial': row.get('remedial', 'False').lower() == 'true',
                            'max_retake': int(row.get('max_retake', 0) or 0),
                            'classroom_mode': row.get('classroom_mode', 'False').lower() == 'true',
                            'shuffle_questions': row.get('shuffle_questions', 'False').lower() == 'true',
                            'activity_instruction': row.get('activity_instruction', '').strip(),
                            'start_time': row.get('start_time', '').strip() or None,
                            'end_time': row.get('end_time', '').strip() or None,
                        },
                        'questions': {}
                    }

                # Add question data if present
                question_order = row.get('question_order', '').strip()
                if question_order and question_order.isdigit():
                    question_order = int(question_order)
                    question_key = f"{activity_key}|{question_order}"

                    if question_key not in activities_data[activity_key]['questions']:
                        activities_data[activity_key]['questions'][question_key] = {
                            'order': question_order,
                            'question_text': row.get('question_text', '').strip(),
                            'correct_answer': row.get('correct_answer', '').strip(),
                            'quiz_type': row.get('quiz_type', '').strip(),
                            'score': float(row.get('question_score', 0) or 0),
                            'choices': []
                        }

                    # Add choice data if present
                    choice_order = row.get('choice_order', '').strip()
                    choice_text = row.get('choice_text', '').strip()
                    is_left_side = row.get('is_left_side', 'False').lower() == 'true'

                    if choice_order and choice_order.isdigit() and choice_text:
                        choice_data = {
                            'order': int(choice_order),
                            'text': choice_text,
                            'is_left_side': is_left_side
                        }
                        if choice_data not in activities_data[activity_key]['questions'][question_key]['choices']:
                            activities_data[activity_key]['questions'][question_key]['choices'].append(choice_data)

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")

        # Process the grouped data
        with transaction.atomic():
            for activity_key, activity_group in activities_data.items():
                try:
                    # Create subject if it doesn't exist
                    subject, created = Subject.objects.get_or_create(
                        subject_name=activity_group['subject_name'],
                        defaults={'subject_code': activity_group['subject_name'][:10].upper()}
                    )

                    # Get or create activity type
                    activity_type_name = activity_group['activity_data']['activity_type']
                    activity_type = None
                    if activity_type_name:
                        activity_type, created = ActivityType.objects.get_or_create(name=activity_type_name)

                    # Prepare activity data
                    activity_data = activity_group['activity_data'].copy()
                    activity_data['activity_type'] = activity_type
                    activity_data['subject'] = subject

                    # Get or create activity
                    activity, created = Activity.objects.get_or_create(
                        activity_name=activity_group['activity_name'],
                        subject=subject,
                        defaults=activity_data
                    )

                    if not created:
                        # Update existing activity
                        for key, value in activity_group['activity_data'].items():
                            setattr(activity, key, value)
                        activity.activity_type = activity_type
                        activity.save()
                        updated_count += 1
                    else:
                        imported_count += 1

                    # Process questions
                    for question_key, question_data in activity_group['questions'].items():
                        # Get or create quiz type
                        quiz_type_name = question_data['quiz_type']
                        quiz_type = None
                        if quiz_type_name:
                            quiz_type, created = QuizType.objects.get_or_create(name=quiz_type_name)

                        # Get or create question
                        question, created = ActivityQuestion.objects.get_or_create(
                            activity=activity,
                            question_text=question_data['question_text'],
                            defaults={
                                'correct_answer': question_data['correct_answer'],
                                'quiz_type': quiz_type,
                                'score': question_data['score']
                            }
                        )

                        # Process choices
                        for choice_data in question_data['choices']:
                            QuestionChoice.objects.get_or_create(
                                question=question,
                                choice_text=choice_data['text'],
                                defaults={'is_left_side': choice_data['is_left_side']}
                            )

                except Exception as e:
                    errors.append(f"Activity {activity_group['activity_name']}: {str(e)}")

        # Display success/error messages
        if errors:
            for error in errors:
                messages.error(request, error)
        
        if imported_count > 0 or updated_count > 0:
            messages.success(request, f'Successfully imported {imported_count} activities and updated {updated_count} activities.')
        else:
            messages.warning(request, 'No activities were imported.')

        return redirect('import_and_export_activity_page')

    except Exception as e:
        messages.error(request, f'Failed to process file: {str(e)}')
        return redirect('import_and_export_activity_page')


# ──────────────────────────────────────────────────────────────────────
# Per-activity question import (Multiple Choice / True-False) + templates
# ──────────────────────────────────────────────────────────────────────
from django.views import View
from django.utils.decorators import method_decorator
from io import TextIOWrapper
from activity.utils.question_import import (
    parse_multiple_choice_csv, parse_true_false_csv,
    parse_multiple_choice_excel, parse_true_false_excel,
)


_MC_TEMPLATE = (
    "question_text,points,choice1,choice2,choice3,choice4,correct_answer\n"
    '"What is 2+2?",,"3","4","5","6","4"\n'
)

_TF_TEMPLATE = (
    "question_text,points,correct_answer\n"
    '"The sky is blue.",,"True"\n'
    '"Water boils at 50C.",,"False"\n'
)

_MC_HEADERS = ["question_text", "points", "choice1", "choice2", "choice3", "choice4", "correct_answer"]
_MC_EXAMPLE = ["What is 2+2?", "", "3", "4", "5", "6", "4"]

_TF_HEADERS = ["question_text", "points", "correct_answer"]
_TF_EXAMPLES = [
    ["The sky is blue.", "", "True"],
    ["Water boils at 50C.", "", "False"],
]


def _build_excel_template(headers, examples):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    header_fill = PatternFill(start_color="217346", end_color="217346", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
    rows = examples if isinstance(examples[0], list) else [examples]
    for row_idx, row_data in enumerate(rows, 2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=val)
    return wb


@method_decorator(login_required, name="dispatch")
class DownloadImportTemplateView(View):
    def get(self, request):
        kind = request.GET.get("type", "mc")
        fmt = request.GET.get("format", "csv")

        if fmt == "xlsx":
            if kind == "tf":
                wb = _build_excel_template(_TF_HEADERS, _TF_EXAMPLES)
                filename = "true_false_template.xlsx"
            else:
                wb = _build_excel_template(_MC_HEADERS, [_MC_EXAMPLE])
                filename = "multiple_choice_template.xlsx"
            import io as _io
            buf = _io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{filename}"'
            return resp

        if kind == "tf":
            body, filename = _TF_TEMPLATE, "true_false_template.csv"
        else:
            body, filename = _MC_TEMPLATE, "multiple_choice_template.csv"
        resp = HttpResponse(body, content_type="text/csv")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp


@method_decorator(login_required, name="dispatch")
class ImportQuestionsUnifiedView(View):
    def post(self, request, activity_id):
        from django.shortcuts import get_object_or_404
        activity = get_object_or_404(Activity, pk=activity_id)
        kind = request.POST.get("import_type", "mc")
        upload = request.FILES.get("csv_file")
        if not upload:
            messages.error(request, "No file uploaded.")
            return redirect("add_quiz_type", activity_id=activity.id)

        filename = upload.name.lower()
        is_excel = filename.endswith(".xlsx") or filename.endswith(".xls")
        default_pts = request.session.get(f'default_points_{activity_id}')

        if is_excel:
            parser = parse_true_false_excel if kind == "tf" else parse_multiple_choice_excel
            rows, errors = parser(upload.file, default_points=default_pts)
        else:
            text = TextIOWrapper(upload.file, encoding="utf-8")
            parser = parse_true_false_csv if kind == "tf" else parse_multiple_choice_csv
            rows, errors = parser(text, default_points=default_pts)

        if errors:
            for err in errors:
                messages.error(request, f"Row {err.row_number}: {err.message}")
            return redirect("add_quiz_type", activity_id=activity.id)

        from activity.utils.score_validation import validate_exact_total
        existing = request.session.get("questions", {}).get(str(activity.id), [])
        projected_sum = sum(float(q.get("score") or 0) for q in existing + rows)
        mismatch = validate_exact_total(activity, projected_sum)
        if mismatch:
            messages.error(request, mismatch.message())
            return redirect("add_quiz_type", activity_id=activity.id)

        session_questions = request.session.get("questions", {})
        session_questions.setdefault(str(activity.id), []).extend(rows)
        request.session["questions"] = session_questions
        request.session.modified = True
        messages.success(request, f"Imported {len(rows)} question(s).")
        return redirect("add_quiz_type", activity_id=activity.id)
