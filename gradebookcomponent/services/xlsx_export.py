"""[Classedge LMS] XLSX export — per-subject gradebook (class record style).

Mirrors `csv_export.build_gradebook_csv` but writes an openpyxl workbook so
teachers get a styled Excel file instead of CSV. Columns:

  Student ID | Last Name | First Name |
  <Activity N> (<max>)  …  |
  <Component N> Subtotal (%)  …  |
  Final Grade (%)
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from accounts.models import CustomUser
from activity.models.activity_model import Activity
from activity.models.student_activity_model import StudentActivity
from gradebookcomponent.models import GradeBookComponents
from gradebookcomponent.services.grades import (
    compute_component_subtotal,
    compute_weighted_grade,
)


def build_gradebook_xlsx(subject, term):
    """Return an in-memory bytes buffer containing the styled workbook."""

    activity_qs = Activity.objects.filter(subject=subject, is_graded=True)
    if term is not None:
        activity_qs = activity_qs.filter(term=term)
    activities = list(activity_qs.order_by("start_time", "local_id"))

    component_qs = GradeBookComponents.objects.filter(subject=subject)
    if term is not None:
        component_qs = component_qs.filter(term=term)
    components = list(component_qs)

    sa_qs = StudentActivity.objects.filter(subject=subject)
    if term is not None:
        sa_qs = sa_qs.filter(term=term)
    student_ids = list(sa_qs.values_list("student_id", flat=True).distinct())
    students = CustomUser.objects.filter(pk__in=student_ids).order_by(
        "last_name", "first_name"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = (subject.subject_name or "Class Record")[:31]

    forest_fill = PatternFill(start_color="1B4332", end_color="1B4332", fill_type="solid")
    gold_fill = PatternFill(start_color="B7925A", end_color="B7925A", fill_type="solid")
    cream_fill = PatternFill(start_color="F4EDE2", end_color="F4EDE2", fill_type="solid")
    header_font = Font(bold=True, color="FAF7F2", name="Calibri", size=11)
    sub_header_font = Font(bold=True, color="1B4332", name="Calibri", size=11)
    body_font = Font(name="Calibri", size=10)
    thin = Side(style="thin", color="BFB7A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Title row
    title_parts = [subject.subject_name or "Class Record"]
    if term is not None:
        title_parts.append(getattr(term, "term_name", ""))
    title = " — ".join(p for p in title_parts if p)
    total_cols = 3 + len(activities) + len(components) + 1
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color="1B4332")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws.row_dimensions[1].height = 24

    # Header row
    header_row = 3
    headers = ["Student ID", "Last Name", "First Name"]
    for a in activities:
        headers.append(f"{a.activity_name}\n({a.max_score})")
    for c in components:
        headers.append(f"{c.gradebook_name or ''}\nSubtotal (%)")
    headers.append("Final Grade (%)")

    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        # Color demographic cols differently from score/component cols.
        if col_idx <= 3:
            cell.fill = forest_fill
        elif col_idx == total_cols:
            cell.fill = gold_fill
        elif col_idx > 3 + len(activities):
            cell.fill = gold_fill
        else:
            cell.fill = forest_fill
    ws.row_dimensions[header_row].height = 38

    # Body rows
    row_idx = header_row + 1
    for student in students:
        last_name = student.last_name or ""
        first_name = student.first_name or ""
        if not last_name and not first_name:
            last_name = student.username or ""

        ident = getattr(student, "id_number", "") or student.id

        sa_map_qs = StudentActivity.objects.filter(
            student=student, subject=subject
        ).prefetch_related("score_logs")
        if term is not None:
            sa_map_qs = sa_map_qs.filter(term=term)
        sa_map = {sa.activity_id: sa for sa in sa_map_qs}

        values = [ident, last_name, first_name]

        for a in activities:
            sa = sa_map.get(a.local_id)
            if sa is None:
                values.append("")
            else:
                has_override = bool(list(sa.score_logs.all())[:1]) if hasattr(sa, "score_logs") else False
                marker = "*" if has_override else ""
                values.append(f"{sa.total_score}{marker}")

        for c in components:
            values.append(compute_component_subtotal(student, subject, term, c))

        values.append(compute_weighted_grade(student, subject, term))

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = left if col_idx <= 3 else center
            if row_idx % 2 == 0:
                cell.fill = cream_fill
            # Highlight final grade column.
            if col_idx == total_cols:
                cell.font = sub_header_font
        row_idx += 1

    # Column widths
    widths = [16, 22, 22] + [14] * len(activities) + [16] * len(components) + [16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes: keep title + header + first 3 demographic cols pinned
    ws.freeze_panes = ws.cell(row=header_row + 1, column=4)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
