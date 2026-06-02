import json
import openpyxl
from decimal import Decimal
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import Sum
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import now
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from gradebookcomponent.models import GradeBookComponents, TermGradeBookComponents, GradeVisibilitySettings, ActivityTypePercentage
from course.models import Semester, Term, Attendance
from subject.models import Subject
from activity.models import Activity, ActivityQuestion, StudentActivity, ActivityType

@login_required
def studentTotalScore(request, student_id, subject_id):
    selected_semester_id = request.GET.get('semester')
    selected_term_id = request.GET.get('term', 'all')

    if selected_semester_id and selected_semester_id != 'None':
        current_semester = get_object_or_404(Semester, id=selected_semester_id)
    else:
        current_semester = Semester.current()
        if not current_semester:
            return render(request, 'gradebookcomponent/activityGrade/studentGrade.html', {
                'error': 'No current semester found.'
            })

    student = get_object_or_404(Subject._meta.apps.get_model('accounts', 'CustomUser'), id=student_id)
    subject = get_object_or_404(Subject, id=subject_id, subjectenrollment__student=student, subjectenrollment__semester=current_semester)

    activity_types = ActivityType.objects.all()
    terms = Term.objects.filter(semester=current_semester)

    term_scores_data = []

    student_activities = StudentActivity.objects.select_related(
        'activity', 'activity__activity_type', 'term', 'activity__subject'
    ).filter(term__semester=current_semester, student=student, activity__subject=subject)

    gradebook_components = GradeBookComponents.objects.filter(term__semester=current_semester).select_related('term', 'subject', 'activity_type')

    term_percentages = {}
    term_gradebook_components = TermGradeBookComponents.objects.filter(
        term__semester=current_semester, subjects=subject
    ).select_related('term')

    for component in term_gradebook_components:
        term_percentages[component.term.id] = float(component.percentage)

    gradebook_lookup = {}
    for component in gradebook_components:
        if component.activity_type is None or component.term is None:
            continue
        gradebook_lookup[(component.subject.subject_name, component.term.term_name, component.activity_type.name)] = float(component.percentage)

    for term in terms:
        if selected_term_id != 'all' and str(term.id) != selected_term_id:
            continue

        student_scores_data = []
        term_has_data = False
        term_total_score = 0
        term_max_score = 0

        for activity_type in activity_types:
            activities = Activity.objects.filter(term=term, activity_type=activity_type, subject=subject, status=True)

            for activity in activities:
                student_activity = student_activities.filter(activity=activity).first()

                if not student_activity:
                    continue

                total_score = student_activity.total_score
                if activity_type.name == "Participation":
                    sa = StudentActivity.objects.filter(student=student, activity=activity).first()
                    total_score = (sa.total_score or 0) if sa else 0
                    max_score = activity.max_score or 0
                else:
                    total_score = student_activity.total_score if student_activity else 0
                    max_score = activity.max_score or 0

                student_scores_data.append({
                    'activity': activity,
                    'activity_type': activity_type.name,
                    'total_score': total_score,
                    'max_score': max_score,
                })

                term_total_score += total_score
                term_max_score += max_score
                term_has_data = True

        if term_has_data:
            term_scores_data.append({
                'term': term.term_name,
                'total_score': term_total_score,
                'max_score': term_max_score,
            })

    final_grade = 0
    total_max_score = 0
    for term_data in term_scores_data:
        final_grade += term_data['total_score']
        total_max_score += term_data['max_score']

    final_percentage = (final_grade / total_max_score) * 100 if total_max_score > 0 else 0

    return render(request, 'gradebookcomponent/activityGrade/studentGrade.html', {
        'term_scores_data': term_scores_data,
        'final_grade': round(final_grade, 2),
        'final_percentage': round(final_percentage, 2),
        'terms': terms,
        'selected_semester_id': selected_semester_id,
        'selected_term_id': selected_term_id,
        'activity_types': activity_types,
        'subject': subject,
        'student': student,
    })


# display total grade
@login_required
def studentTotalScoreForActivityType(request):
    is_teacher = request.user.is_teacher
    students = Subject._meta.apps.get_model('accounts', 'CustomUser').objects.filter(profile__role__name__iexact='student')
    current_semester = Semester.objects.filter(start_date__lte=timezone.now(), end_date__gte=timezone.now()).first()

    selected_student_id = request.GET.get('student') if is_teacher else request.user.id
    selected_subject_id = request.GET.get('subject')
    selected_term_id = request.GET.get('term')

    selected_student = None
    if selected_student_id:
        selected_student = get_object_or_404(Subject._meta.apps.get_model('accounts', 'CustomUser'), id=selected_student_id)

    subjects = (
        Subject.objects.filter(subjectenrollment__semester=current_semester)
        .exclude(is_coil=True).exclude(is_hali=True)
        .distinct()
    )

    terms = Term.objects.filter(semester=current_semester) if selected_subject_id else []

    return render(request, 'gradebookcomponent/activityGrade/studentScore.html', {
        'students': students,
        'subjects': subjects,
        'terms': terms,
        'selected_student': selected_student,
        'selected_subject_id': selected_subject_id,
        'selected_term_id': selected_term_id,
        'is_teacher': is_teacher,
    })


@login_required
def getSemesters(request):
    now_local = timezone.localtime(timezone.now())
    semesters = Semester.objects.all().order_by('-start_date')

    data = [
        {
            'id': semester.id,
            'name': semester.semester_name,
            'start_date': semester.start_date.strftime('%B %d, %Y'),
            'end_date': semester.end_date.strftime('%B %d, %Y'),
            'is_current': semester.start_date <= now_local.date() <= semester.end_date
        }
        for semester in semesters
    ]

    return JsonResponse({'semesters': data})


# fetch subject
@login_required
def getSubjects(request):
    semester_id = request.GET.get('semester_id')

    if not semester_id:
        return JsonResponse({'error': 'Semester ID is required'}, status=400)

    try:
        semester = Semester.objects.get(id=semester_id)
    except Semester.DoesNotExist:
        return JsonResponse({'error': 'Invalid semester ID'}, status=400)

    subjects = (
        Subject.objects.filter(subjectenrollment__semester=semester)
        .exclude(is_coil=True).exclude(is_hali=True)
        .distinct()
    )

    data = [
        {
            'id': subject.id,
            'name': subject.subject_name,
        }
        for subject in subjects
    ]

    return JsonResponse({'subjects': data})


# Teacher (allow student to see grade)
@login_required
def allowGradeVisibility(request, student_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_id = data.get('subject_id')
            term_id = data.get('term_id')
            is_visible = data.get('is_visible')

            if subject_id is None:
                return JsonResponse({'status': 'failure', 'message': 'Missing subject ID.'}, status=400)

            subject = get_object_or_404(Subject, id=subject_id)
            term = None if term_id is None else get_object_or_404(Term, id=term_id)

            setting, created = GradeVisibilitySettings.objects.update_or_create(
                teacher=request.user,
                subject=subject,
                term=term,
                defaults={'is_visible': is_visible}
            )

            return JsonResponse({'status': 'success', 'message': 'Grade visibility updated successfully.'})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'failure', 'message': 'Invalid JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'failure', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'failure', 'message': 'Invalid request method.'}, status=405)

@login_required
def student_grades(request):
    semesters = Semester.objects.all()
    current_semester = Semester.objects.filter(
        start_date__lte=now(),
        end_date__gte=now()
    ).first()

    user_role = (request.user.role_name or 'unknown')

    if user_role == 'student':
        return _student_my_grades(request, semesters, current_semester)

    return teacher_student_grades_summary(request)


def _student_my_grades(request, semesters, current_semester):
    """Dedicated student view: enrolled subjects with the proper weighted
    gradebook calculation, term breakdown, and per-activity scores. Uses the
    same `get_student_activity_summary()` helper that the registrar/teacher
    gradebook uses, so the numbers students see here match the official record.
    """
    from course.models import SubjectEnrollment, Term, Attendance
    from activity.utils.grade_calculation_utils import get_student_activity_summary

    selected_id = request.GET.get('semester')
    if selected_id:
        try:
            selected_semester = Semester.objects.filter(pk=int(selected_id)).first() or current_semester
        except (TypeError, ValueError):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    enrollments = SubjectEnrollment.objects.filter(
        student=request.user, status='enrolled'
    ).select_related('subject', 'subject__assign_teacher', 'semester').exclude(
        subject__is_coil=True
    ).exclude(subject__is_hali=True)
    if selected_semester:
        enrollments = enrollments.filter(semester=selected_semester)

    passing_grade = float(selected_semester.passing_grade) if selected_semester else 75.0

    profile = getattr(request.user, 'profile', None)
    student_name = (
        f"{profile.first_name} {profile.last_name}".strip()
        if profile and profile.first_name else ''
    )

    rows = []
    term_columns = []
    if selected_semester:
        terms = list(Term.objects.filter(semester=selected_semester).order_by('start_date'))
        term_columns = [t.term_name for t in terms]
        for enr in enrollments:
            subject = enr.subject
            if not subject:
                continue
            activities_qs = StudentActivity.objects.select_related(
                'activity', 'activity__activity_type', 'activity__subject', 'term', 'student__profile'
            ).filter(
                term__semester=selected_semester,
                activity__subject=subject,
                activity__status=True,
            )
            attendance_qs = Attendance.objects.select_related('subject', 'student').filter(
                subject=subject,
                graded=True,
                date__range=(selected_semester.start_date, selected_semester.end_date),
            )
            summary = get_student_activity_summary(
                selected_semester, subject, terms, activities_qs, attendance_qs, request.user
            )
            # The helper filters to request.user when the caller is a student,
            # so the summary holds at most one entry. Look up by name first
            # (matches the API), then fall back to the single value to be
            # robust against name-key whitespace drift.
            data = summary.get(student_name) if student_name else None
            if not data and summary:
                data = next(iter(summary.values()), None)
            if not data:
                data = {
                    'final_grade': None,
                    'term_grades': {},
                    'activities': {},
                    'attendance': {},
                }

            final_grade = data.get('final_grade')
            try:
                final_grade_f = float(final_grade) if final_grade is not None else None
            except (TypeError, ValueError):
                final_grade_f = None

            term_breakdown = []
            term_grades_map = data.get('term_grades') or {}
            for term_name in term_columns:
                term_data = term_grades_map.get(term_name) or {}
                term_breakdown.append({
                    'name': term_name,
                    'total': term_data.get('total_grade'),
                    'visibility': term_data.get('visibility', True),
                    'categories': term_data.get('categories', []),
                })

            items = []
            for term_name, types in (data.get('activities') or {}).items():
                for type_name, type_data in types.items():
                    for breakdown in type_data.get('activities_breakdown', []):
                        max_s = breakdown.get('max_score') or 0
                        score_s = breakdown.get('student_score') or 0
                        items.append({
                            'name': breakdown.get('activity_name'),
                            'term_name': term_name,
                            'type_name': type_name,
                            'score': score_s,
                            'max_score': max_s,
                            'pct': (score_s / max_s * 100) if max_s else 0,
                        })

            rows.append({
                'subject': subject,
                'final_grade': final_grade_f,
                'is_passing': (final_grade_f is not None and final_grade_f >= passing_grade),
                'terms': term_breakdown,
                'items': items,
                'has_activities': bool(items),
            })

    return render(request, 'grades/my-grades.html', {
        'semesters': semesters,
        'selected_semester': selected_semester,
        'rows': rows,
        'term_columns': term_columns,
        'passing_grade': passing_grade,
    })


@login_required
def manage_grade_visibility(request):
    """
    View for teachers/admins to manage grade visibility settings per subject and (optionally) per term.
    """
    today = timezone.now().date()
    current_semester = Semester.objects.filter(start_date__lte=today, end_date__gte=today).first()

    # Get subjects taught by the teacher in the current semester; admins see all subjects
    if request.user.is_teacher:
        subjects = Subject.objects.filter(
            subjectenrollment__teacher=request.user,
            subjectenrollment__semester=current_semester
        ).distinct()
    else:
        subjects = Subject.objects.filter(subjectenrollment__semester=current_semester).distinct()

    # Terms for current semester
    terms = Term.objects.filter(semester=current_semester).order_by('term_name')

    # Handle form submission
    if request.method == 'POST':
        subject_id = request.POST.get('subject')
        term_id = request.POST.get('term')
        is_visible = request.POST.get('is_visible') == 'on'

        subject = get_object_or_404(Subject, id=subject_id)

        if not term_id:
            # Subject-level setting (no specific term)
            GradeVisibilitySettings.objects.update_or_create(
                teacher=request.user,
                subject=subject,
                term=None,
                defaults={'is_visible': is_visible}
            )
            messages.success(request, f'Grade visibility for {subject.subject_name} updated successfully.')
        else:
            term = get_object_or_404(Term, id=term_id)
            GradeVisibilitySettings.objects.update_or_create(
                teacher=request.user,
                subject=subject,
                term=term,
                defaults={'is_visible': is_visible}
            )
            messages.success(request, f'Grade visibility for {subject.subject_name} in {term.term_name} updated successfully.')

        return redirect('manage_grade_visibility')

    # Existing settings for the teacher/admin
    visibility_settings = GradeVisibilitySettings.objects.filter(teacher=request.user)

    # Organize settings by subject and term
    settings_dict = {}
    for setting in visibility_settings:
        sid = setting.subject.id
        if sid not in settings_dict:
            settings_dict[sid] = {'subject': setting.subject, 'terms': {}, 'subject_level': None}
        if setting.term:
            settings_dict[sid]['terms'][setting.term.id] = setting.is_visible
        else:
            settings_dict[sid]['subject_level'] = setting.is_visible

    context = {
        'subjects': subjects,
        'terms': terms,
        'settings': settings_dict,
        'current_semester': current_semester,
    }

    return render(request, 'gradebookcomponent/grade_visibility/manage_grade_visibility.html', context)

@csrf_exempt
def toggle_grade_visibility(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            subject_id = data.get('subject_id')
            term_id = data.get('term_id')
            is_visible = data.get('is_visible')
            delete = data.get('delete', False)

            if subject_id is None:
                return JsonResponse({'status': 'failure', 'message': 'Missing subject ID.'}, status=400)

            user_role = request.user.role_name
            if user_role not in ['teacher', 'admin']:
                return JsonResponse({'status': 'failure', 'message': 'Permission denied.'}, status=403)

            subject = get_object_or_404(Subject, id=subject_id)
            term = None if term_id is None else get_object_or_404(Term, id=term_id)

            if delete:
                try:
                    setting = GradeVisibilitySettings.objects.get(
                        teacher=request.user,
                        subject=subject,
                        term=term
                    )
                    setting.delete()
                    return JsonResponse({'status': 'success', 'message': 'Grade visibility setting deleted successfully.'})
                except GradeVisibilitySettings.DoesNotExist:
                    return JsonResponse({'status': 'failure', 'message': 'Setting not found.'}, status=404)

            if is_visible is None:
                return JsonResponse({'status': 'failure', 'message': 'Missing visibility status.'}, status=400)

            setting, created = GradeVisibilitySettings.objects.update_or_create(
                teacher=request.user,
                subject=subject,
                term=term,
                defaults={'is_visible': is_visible}
            )

            # Notify enrolled students when grades are *made visible* — a clean,
            # unambiguous "grades are now available" event. Deduped per
            # subject+term so re-toggling within 12h doesn't spam the bell.
            if is_visible:
                try:
                    from django.urls import reverse
                    from logs.notifications import notify
                    from course.models import SubjectEnrollment
                    term_label = f" ({term.term_name})" if term else ""
                    grades_path = reverse('my-grades')
                    enrolled = SubjectEnrollment.objects.filter(
                        subject=subject, status='enrolled',
                    ).select_related('student')
                    for enr in enrolled:
                        notify(
                            enr.student,
                            f"Your grades for {subject.subject_name}{term_label} are now available.",
                            name="Grades posted",
                            path=grades_path,
                            entity_type="grade_visible",
                            entity_id=f"{subject.id}:{term_id or ''}",
                            created_by=request.user,
                            dedupe_hours=12,
                        )
                except Exception:
                    pass  # notifications are best-effort; never block the toggle

            action = "shown to" if is_visible else "hidden from"
            return JsonResponse({'status': 'success', 'message': f'Grades are now {action} students.', 'is_visible': setting.is_visible})

        except json.JSONDecodeError:
            return JsonResponse({'status': 'failure', 'message': 'Invalid JSON.'}, status=400)
        except Exception as e:
            return JsonResponse({'status': 'failure', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'failure', 'message': 'Method not allowed.'}, status=405)


@login_required
def export_grades_excel(request):
    semester_id = request.GET.get('semester')
    subject_id = request.GET.get('subject')

    if not semester_id or not subject_id:
        return JsonResponse({'error': 'Semester and subject are required'}, status=400)

    try:
        semester = Semester.objects.get(id=semester_id)
        subject = Subject.objects.get(id=subject_id)
    except (Semester.DoesNotExist, Subject.DoesNotExist):
        return JsonResponse({'error': 'Invalid semester or subject'}, status=400)

    data = get_student_activity_summary(request, semester_id=semester_id, subject_id=subject_id)

    if not data:
        return JsonResponse({'error': 'No data available for export'}, status=404)

    return _export_grades_excel_dynamic(semester, subject, data)


def _export_grades_excel_dynamic(semester, subject, data):
    """Category-driven Excel export.

    The layout is derived from the actual GradeBookComponents (categories) the
    teacher created for each term. Each category becomes one block:
        [sub_activity_1 instances ...] [sub_total] ... [Category Total]
    followed by a per-term "Term Grade" column. Final Grade and Transmuted
    columns sit at the right.
    """

    def format_percentage(decimal_value):
        try:
            value = float(decimal_value)
            if value <= 1:
                return f"{int(value * 100)}%"
            else:
                return f"{int(value)}%"
        except (TypeError, ValueError):
            return str(decimal_value) or '--'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Grades"

    header_font = Font(bold=True, color="FFFFFF")

    primary_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    success_fill = PatternFill(start_color="198754", end_color="198754", fill_type="solid")
    info_fill = PatternFill(start_color="0DCAF0", end_color="0DCAF0", fill_type="solid")
    warning_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    danger_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
    secondary_fill = PatternFill(start_color="6C757D", end_color="6C757D", fill_type="solid")

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    any_student_data = next(iter(data.values()), None)
    terms = list((any_student_data or {}).get('term_grades', {}).keys()) if any_student_data else []

    if not terms:
        return JsonResponse({'error': 'No terms found for export'}, status=404)

    # ---- Build per-term layout from the actual GradeBookComponents ----
    term_id_by_name = {}
    for term_obj in Term.objects.filter(semester=semester, term_name__in=terms):
        term_id_by_name[term_obj.term_name] = term_obj.id

    def _instance_count_for(activity_type_name):
        # Count across all students so columns line up even if some students
        # are missing entries for a given activity type.
        max_count = 0
        is_attendance = activity_type_name.lower() == 'attendance'
        for student_data in data.values():
            for term_name in terms:
                if is_attendance:
                    records = (
                        (student_data.get('attendance') or {})
                        .get(term_name, {})
                        .get('records') or []
                    )
                    max_count = max(max_count, len(records))
                else:
                    breakdown = (
                        (student_data.get('activities') or {})
                        .get(term_name, {})
                        .get(activity_type_name, {})
                        .get('activities_breakdown') or []
                    )
                    max_count = max(max_count, len(breakdown))
        return max(max_count, 1)

    layouts = {}  # term_name -> list of categories with their sub-activities
    for term_name in terms:
        term_id = term_id_by_name.get(term_name)
        if term_id is None:
            layouts[term_name] = []
            continue
        components = (
            GradeBookComponents.objects
            .filter(subject=subject, term_id=term_id)
            .prefetch_related('activity_type_percentages__activity_type')
            .order_by('id')
        )
        term_layout = []
        for component in components:
            sub_activities = []
            for atp in component.activity_type_percentages.all():
                type_name = atp.activity_type.name
                sub_activities.append({
                    'type_name': type_name,
                    'sub_pct': atp.percentage,
                    'is_attendance': type_name.lower() == 'attendance',
                    'instance_count': _instance_count_for(type_name),
                })
            cols = sum(s['instance_count'] + 1 for s in sub_activities) + 1  # + Category Total
            term_layout.append({
                'category': component.gradebook_category,
                'percentage': component.percentage,
                'sub_activities': sub_activities,
                'cols': cols,
            })
        layouts[term_name] = term_layout

    # Pick a deterministic palette for category headers. Cycles through the
    # original set so visual cues stay similar to the legacy export.
    category_palette = [info_fill, warning_fill, danger_fill, success_fill, primary_fill]

    # ----------------------------- Helpers ----------------------------------

    def get_activity_count(student_data, term, activity_type):
        if (
            not student_data
            or not student_data.get('activities')
            or not student_data['activities'].get(term)
            or not student_data['activities'][term].get(activity_type)
        ):
            return 1
        breakdown = student_data['activities'][term][activity_type].get('activities_breakdown') or []
        return max(len(breakdown), 1)

    def get_activity_name(student_data, term, activity_type, index):
        if (
            not student_data
            or not student_data.get('activities')
            or not student_data['activities'].get(term)
            or not student_data['activities'][term].get(activity_type)
        ):
            return None
        breakdown = student_data['activities'][term][activity_type].get('activities_breakdown') or []
        if len(breakdown) <= index:
            return None
        return breakdown[index].get('activity_name')

    def get_activity_score(student_data, term, activity_type, index):
        if (
            not student_data
            or not student_data.get('activities')
            or not student_data['activities'].get(term)
            or not student_data['activities'][term].get(activity_type)
        ):
            return '--'
        breakdown = student_data['activities'][term][activity_type].get('activities_breakdown') or []
        if len(breakdown) <= index:
            return '--'
        score = breakdown[index].get('student_score', breakdown[index].get('score'))
        return score if score is not None else '--'

    def get_activity_type_total(student_data, term, activity_type):
        if (
            not student_data
            or not student_data.get('activities')
            or not student_data['activities'].get(term)
            or not student_data['activities'][term].get(activity_type)
        ):
            return '--'
        total = student_data['activities'][term][activity_type].get('total_score')
        return total if total is not None else '--'

    def get_attendance_record(student_data, term, index):
        if (
            not student_data
            or not student_data.get('attendance')
            or not student_data['attendance'].get(term)
        ):
            return '--'
        records = student_data['attendance'][term].get('records') or []
        if len(records) <= index:
            return '--'
        return records[index].get('status') or '--'

    def get_attendance_total(student_data, term):
        if (
            not student_data
            or not student_data.get('attendance')
            or not student_data['attendance'].get(term)
        ):
            return '--'
        total = student_data['attendance'][term].get('total_attendance')
        return total if total is not None else '--'

    def get_category_total(student_data, term, category):
        if (
            not student_data
            or not student_data.get('term_grades')
            or not student_data['term_grades'].get(term)
        ):
            return '--'
        value = student_data['term_grades'][term].get(category)
        return format_percentage(value) if value is not None else '--'

    def get_term_grade(student_data, term):
        if (
            not student_data
            or not student_data.get('term_grades')
            or not student_data['term_grades'].get(term)
        ):
            return '--'
        grade = student_data['term_grades'][term].get('total_grade')
        return grade if grade is not None else '--'

    def get_final_grade(student_data):
        if not student_data or student_data.get('final_grade') is None:
            return '--'
        return student_data['final_grade']

    def get_transmuted_grade(student_data):
        return student_data.get('transmuted_grade', '--') if student_data else '--'

    # ----------------------------- Headers ----------------------------------

    # Row 1: per-term super-header + Final Grade / Transmuted columns.
    ws.cell(row=1, column=1, value="Student Name").font = header_font
    ws.cell(row=1, column=1).fill = success_fill
    ws.cell(row=1, column=1).alignment = center_alignment
    ws.cell(row=1, column=1).border = thin_border
    ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)

    col_num = 2
    for term in terms:
        term_cols = sum(c['cols'] for c in layouts[term]) + 1  # +1 for Term Grade
        if term_cols == 1:
            term_cols = 1
        cell = ws.cell(row=1, column=col_num, value=f"{term} Components")
        cell.font = header_font
        cell.fill = primary_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=col_num, end_row=1, end_column=col_num + term_cols - 1)
        col_num += term_cols

    for header in ["Final Grade", "Transmuted"]:
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = secondary_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.merge_cells(start_row=1, start_column=col_num, end_row=3, end_column=col_num)
        col_num += 1

    # Row 2: per-category headers + per-term "Term Grade" column header.
    col_num = 2
    for term in terms:
        for cat_idx, category in enumerate(layouts[term]):
            fill = category_palette[cat_idx % len(category_palette)]
            cell = ws.cell(
                row=2,
                column=col_num,
                value=f"{category['category']} {format_percentage(category['percentage'])}",
            )
            cell.font = header_font
            cell.fill = fill
            cell.alignment = center_alignment
            cell.border = thin_border
            ws.merge_cells(
                start_row=2,
                start_column=col_num,
                end_row=2,
                end_column=col_num + category['cols'] - 1,
            )
            col_num += category['cols']

        cell = ws.cell(row=2, column=col_num, value="Term Grade")
        cell.font = header_font
        cell.fill = primary_fill
        cell.alignment = center_alignment
        cell.border = thin_border
        ws.merge_cells(start_row=2, start_column=col_num, end_row=3, end_column=col_num)
        col_num += 1

    # Row 3: per-sub-activity instance / total / category-total column headers.
    col_num = 2
    for term in terms:
        for cat_idx, category in enumerate(layouts[term]):
            fill = category_palette[cat_idx % len(category_palette)]
            for sub in category['sub_activities']:
                for i in range(sub['instance_count']):
                    if sub['is_attendance']:
                        label = f"Record {i+1}"
                    else:
                        label = (
                            get_activity_name(any_student_data, term, sub['type_name'], i)
                            or f"{sub['type_name']} {i+1}"
                        )
                    cell = ws.cell(row=3, column=col_num, value=label)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = fill
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    col_num += 1
                cell = ws.cell(row=3, column=col_num, value=f"{sub['type_name']} Total")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = center_alignment
                cell.border = thin_border
                col_num += 1
            cell = ws.cell(row=3, column=col_num, value="Category Total")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = center_alignment
            cell.border = thin_border
            col_num += 1
        # Skip Term Grade column — already merged from row 2.
        col_num += 1

    # ----------------------------- Student rows ------------------------------

    soft_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
    cat_total_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    term_grade_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    grand_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    row_num = 5
    for student_name, student_data in data.items():
        col_num = 1
        ws.cell(row=row_num, column=col_num, value=student_name).border = thin_border
        ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal='left', vertical='center')
        col_num += 1

        for term in terms:
            for category in layouts[term]:
                for sub in category['sub_activities']:
                    for i in range(sub['instance_count']):
                        if sub['is_attendance']:
                            score = get_attendance_record(student_data, term, i)
                        else:
                            score = get_activity_score(student_data, term, sub['type_name'], i)
                        cell = ws.cell(row=row_num, column=col_num, value=score)
                        cell.border = thin_border
                        cell.alignment = center_alignment
                        col_num += 1
                    if sub['is_attendance']:
                        sub_total = get_attendance_total(student_data, term)
                    else:
                        sub_total = get_activity_type_total(student_data, term, sub['type_name'])
                    cell = ws.cell(row=row_num, column=col_num, value=sub_total)
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    cell.fill = soft_fill
                    col_num += 1
                # Category total
                cat_total = get_category_total(student_data, term, category['category'])
                cell = ws.cell(row=row_num, column=col_num, value=cat_total)
                cell.border = thin_border
                cell.alignment = center_alignment
                cell.fill = cat_total_fill
                cell.font = Font(bold=True)
                col_num += 1

            # Term grade
            term_grade = get_term_grade(student_data, term)
            cell = ws.cell(row=row_num, column=col_num, value=term_grade)
            cell.border = thin_border
            cell.alignment = center_alignment
            cell.fill = term_grade_fill
            cell.font = Font(bold=True)
            col_num += 1

        # Final grade + transmuted
        cell = ws.cell(row=row_num, column=col_num, value=get_final_grade(student_data))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = grand_fill
        cell.font = Font(bold=True)
        col_num += 1

        cell = ws.cell(row=row_num, column=col_num, value=get_transmuted_grade(student_data))
        cell.border = thin_border
        cell.alignment = center_alignment
        cell.fill = grand_fill
        cell.font = Font(bold=True)
        row_num += 1

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{subject.subject_name}_{semester.semester_name}_Grades.xlsx"'
    wb.save(response)
    return response


@login_required
def export_my_grades_excel(request):
    """Student-facing grade record export.

    Generates a single-student transcript for one subject in one semester:
    every assessment listed under its term + category, with score, max
    score, and percentage, plus per-term subtotals and a final-grade row.
    Numbers come from the same canonical helper that powers /my-grades/.
    """
    from course.models import SubjectEnrollment, Term as TermModel
    from activity.utils.grade_calculation_utils import (
        get_student_activity_summary as canonical_summary,
    )

    student = request.user
    if (getattr(student, 'role_name', '') or '').lower() == 'student':
        return JsonResponse({'error': 'Export is not available for students.'}, status=403)
    semester_id = request.GET.get('semester')
    subject_id = request.GET.get('subject')
    if not semester_id or not subject_id:
        return JsonResponse({'error': 'Semester and subject are required'}, status=400)

    try:
        semester = Semester.objects.get(id=semester_id)
        subject = Subject.objects.get(id=subject_id)
    except (Semester.DoesNotExist, Subject.DoesNotExist):
        return JsonResponse({'error': 'Invalid semester or subject'}, status=400)

    if subject.is_coil or subject.is_hali:
        return JsonResponse(
            {'error': 'COIL and HALI courses use a separate grading flow and cannot be exported here.'},
            status=400,
        )

    # Enrollment check — students can only export their own subjects.
    if not SubjectEnrollment.objects.filter(
        student=student, subject=subject, semester=semester, status='enrolled'
    ).exists():
        return JsonResponse({'error': 'You are not enrolled in this subject.'}, status=403)

    terms = list(TermModel.objects.filter(semester=semester).order_by('start_date'))
    activities_qs = StudentActivity.objects.select_related(
        'activity', 'activity__activity_type', 'activity__subject', 'term', 'student__profile'
    ).filter(
        term__semester=semester,
        activity__subject=subject,
        activity__status=True,
    )
    attendance_qs = Attendance.objects.select_related('subject', 'student').filter(
        subject=subject,
        graded=True,
        date__range=(semester.start_date, semester.end_date),
    )
    summary = canonical_summary(semester, subject, terms, activities_qs, attendance_qs, student)

    profile = getattr(student, 'profile', None)
    student_name = (
        f"{profile.first_name} {profile.last_name}".strip()
        if profile and profile.first_name else student.get_full_name() or student.username
    )
    data = summary.get(student_name) if student_name else None
    if not data and summary:
        data = next(iter(summary.values()), None)
    data = data or {'final_grade': None, 'term_grades': {}, 'activities': {}}

    return _write_student_grade_workbook(
        student=student,
        student_name=student_name,
        subject=subject,
        semester=semester,
        terms=terms,
        data=data,
    )


def _write_student_grade_workbook(*, student, student_name, subject, semester, terms, data):
    """Build the xlsx response for a single student's grade record.

    Layout:
        ┌─────────────────────────────────────────────────────────┐
        │  GRADE RECORD                                            │
        │  Student: …    Subject: …    Semester: …                 │
        ├──────────────────────────────────────────────────────────┤
        │  Term  │ Category │ Assessment │ Score │ Max │ %  │ Stat │
        │  ─── one row per assessment, grouped by term/category ── │
        │  …                              Term Total: x / y (z%)   │
        │  …                                                       │
        │                                Final Grade: nn.n         │
        └──────────────────────────────────────────────────────────┘
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{subject.subject_short_name or subject.subject_name}"[:31]

    thin = Side(border_style='thin', color='B0B0B0')
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    forest_fill = PatternFill(start_color='1F3527', end_color='1F3527', fill_type='solid')
    gold_fill = PatternFill(start_color='B48A3C', end_color='B48A3C', fill_type='solid')
    cream_fill = PatternFill(start_color='FAF7F2', end_color='FAF7F2', fill_type='solid')
    subtotal_fill = PatternFill(start_color='F0EAD6', end_color='F0EAD6', fill_type='solid')

    headers = ['#', 'Term', 'Category', 'Assessment', 'Score', 'Max Score', '%', 'Status']
    widths = [5, 18, 18, 38, 10, 10, 10, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = w

    # Title block
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title = ws.cell(row=1, column=1, value='Student Grade Record')
    title.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    title.alignment = center
    title.fill = forest_fill
    ws.row_dimensions[1].height = 28

    def _kv(row, label, value):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, color='1F3527')
        ws.cell(row=row, column=1).alignment = left
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(headers))
        v = ws.cell(row=row, column=2, value=value)
        v.alignment = left
        v.font = Font(color='1F3527')

    _kv(2, 'Student', student_name)
    _kv(3, 'Student ID', getattr(profile := getattr(student, 'profile', None), 'id_number', '') or '')
    _kv(4, 'Subject', f"{subject.subject_name}"
        + (f" ({subject.subject_code})" if subject.subject_code else ''))
    _kv(5, 'Teacher', subject.assign_teacher.get_full_name() if subject.assign_teacher else '—')
    _kv(6, 'Semester', f"{semester.semester_name} · {semester.get_academic_year()}")
    _kv(7, 'Generated', timezone.localtime(timezone.now()).strftime('%b %d, %Y %I:%M %p'))

    # Column headers
    header_row = 9
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = center
        cell.fill = gold_fill
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    row = header_row + 1
    item_index = 1
    activities_by_term = data.get('activities') or {}
    term_grades_map = data.get('term_grades') or {}

    for term in terms:
        term_name = term.term_name
        term_types = activities_by_term.get(term_name, {}) or {}
        # Skip empty terms entirely, but show a placeholder so the
        # student sees that the term exists.
        if not term_types:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
            placeholder = ws.cell(row=row, column=1, value=f"{term_name} — no assessments yet")
            placeholder.alignment = center
            placeholder.font = Font(italic=True, color='888888')
            placeholder.fill = cream_fill
            placeholder.border = border
            row += 1
            continue

        for type_name, type_data in term_types.items():
            breakdown = type_data.get('activities_breakdown', []) or []
            for entry in breakdown:
                score = entry.get('student_score') or 0
                max_score = entry.get('max_score') or 0
                pct = (score / max_score * 100) if max_score else 0
                status = 'Submitted' if score > 0 else ('Missed' if max_score else '—')

                values = [
                    item_index,
                    term_name,
                    type_name,
                    entry.get('activity_name') or '—',
                    score,
                    max_score,
                    f"{pct:.1f}%",
                    status,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    cell.border = border
                    cell.alignment = center if col_idx != 4 else left
                item_index += 1
                row += 1

        # Per-term subtotal row
        term_summary = term_grades_map.get(term_name, {}) or {}
        term_total = term_summary.get('total_grade')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        sub_label = ws.cell(row=row, column=1, value=f"{term_name} — Term Grade")
        sub_label.font = Font(bold=True, color='1F3527')
        sub_label.alignment = left
        sub_label.fill = subtotal_fill
        sub_label.border = border
        for col_idx in range(5, 8):
            c = ws.cell(row=row, column=col_idx, value='')
            c.fill = subtotal_fill
            c.border = border
        total_cell = ws.cell(
            row=row, column=8,
            value=(f"{float(term_total):.1f}" if term_total is not None else '—'),
        )
        total_cell.font = Font(bold=True, color='1F3527')
        total_cell.alignment = center
        total_cell.fill = subtotal_fill
        total_cell.border = border
        row += 1

    # Final grade row
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    label = ws.cell(row=row, column=1, value='Final Grade')
    label.font = Font(bold=True, size=12, color='FFFFFF')
    label.alignment = Alignment(horizontal='right', vertical='center')
    label.fill = forest_fill
    label.border = border
    final = data.get('final_grade')
    try:
        final_f = float(final) if final is not None else None
    except (TypeError, ValueError):
        final_f = None
    final_cell = ws.cell(
        row=row, column=8,
        value=(f"{final_f:.1f}" if final_f is not None else '—'),
    )
    final_cell.font = Font(bold=True, size=12, color='FFFFFF')
    final_cell.alignment = center
    final_cell.fill = forest_fill
    final_cell.border = border
    ws.row_dimensions[row].height = 24

    # Passing status footer
    passing_grade = float(getattr(semester, 'passing_grade', 75) or 75)
    if final_f is not None:
        row += 1
        is_passing = final_f >= passing_grade
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        status_cell = ws.cell(
            row=row, column=1,
            value=(
                f"Status: {'PASSED' if is_passing else 'FAILED'} "
                f"(passing grade: {passing_grade:.0f})"
            ),
        )
        status_cell.font = Font(bold=True, color='1F3527' if is_passing else 'B22222')
        status_cell.alignment = center
        status_cell.fill = cream_fill
        status_cell.border = border

    # Freeze the header so it stays visible while scrolling.
    ws.freeze_panes = f'A{header_row + 1}'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_subject = ''.join(c for c in subject.subject_name if c.isalnum() or c in ('_', '-', ' ')).strip().replace(' ', '_')
    safe_semester = ''.join(c for c in semester.semester_name if c.isalnum() or c in ('_', '-', ' ')).strip().replace(' ', '_')
    filename = f"GradeRecord_{safe_subject}_{safe_semester}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# Helper for excel / summaries used above
@login_required
def get_student_activity_summary(request, semester_id, subject_id):
    student = request.user
    subject = get_object_or_404(Subject, id=subject_id, subjectenrollment__student=student, subjectenrollment__semester__id=semester_id)

    activity_types = ActivityType.objects.all()
    terms = Term.objects.filter(semester__id=semester_id)

    term_scores_data = []

    student_activities = StudentActivity.objects.select_related(
        'activity', 'activity__activity_type', 'term', 'activity__subject'
    ).filter(term__semester__id=semester_id, student=student, activity__subject=subject)

    gradebook_components = GradeBookComponents.objects.filter(term__semester__id=semester_id).select_related('term', 'subject', 'activity_type')

    term_percentages = {}
    term_gradebook_components = TermGradeBookComponents.objects.filter(term__semester__id=semester_id, subjects=subject).select_related('term')

    for component in term_gradebook_components:
        term_percentages[component.term.id] = float(component.percentage)

    gradebook_lookup = {}
    for component in gradebook_components:
        if component.activity_type is None or component.term is None:
            continue
        gradebook_lookup[(component.subject.subject_name, component.term.term_name, component.activity_type.name)] = float(component.percentage)

    for term in terms:
        student_scores_data = []
        term_has_data = False
        term_total_score = 0
        term_max_score = 0

        for activity_type in activity_types:
            activities = Activity.objects.filter(term=term, activity_type=activity_type, subject=subject, status=True)

            for activity in activities:
                student_activity = student_activities.filter(activity=activity).first()
                if not student_activity:
                    continue

                total_score = student_activity.total_score
                if activity_type.name == "Participation":
                    sa = StudentActivity.objects.filter(student=student, activity=activity).first()
                    total_score = (sa.total_score or 0) if sa else 0
                    max_score = activity.max_score or 0
                else:
                    total_score = student_activity.total_score if student_activity else 0
                    max_score = ActivityQuestion.objects.filter(activity=activity).aggregate(total_max_score=Sum('score'))['total_max_score'] or 0

                term_max_score += max_score
                percentage = (total_score / max_score) * 100 if max_score > 0 else 0
                status = 'Completed' if total_score > 0 else 'Missed'

                student_scores_data.append({
                    'activity_name': activity.activity_name,
                    'activity_type': activity.activity_type,
                    'is_remedial': activity.remedial,
                    'total_score': total_score,
                    'max_score': max_score,
                    'percentage': round(percentage, 2),
                    'status': status
                })

                term_total_score += total_score
                term_has_data = True

        if term_has_data:
            term_percentage = term_percentages.get(term.id, 0)
            weighted_score = (term_total_score * term_percentage) / 100

            term_scores_data.append({
                'term': term,
                'subject': subject,
                'student_scores_data': student_scores_data,
                'term_total_score': term_total_score,
                'term_max_score': term_max_score,
                'term_percentage': term_percentage,
                'weighted_score': round(weighted_score, 2),
            })

    return {
        student.id: {
            'term_grades': {
                term.term_name: {
                    'activities': {
                        activity_type.name: {
                            'activities_breakdown': [
                                {
                                    'activity_name': activity.activity_name,
                                    'total_score': student_activity.total_score,
                                    'max_score': max_score,
                                    'percentage': round((student_activity.total_score / max_score) * 100, 2) if max_score > 0 else 0,
                                } for activity, student_activity in zip(activities, student_activities.filter(activity__in=activities))
                            ],
                            'total_score': sum(activity.total_score for activity in student_activities.filter(activity__in=activities)),
                            'max_score': sum(max_score for max_score in [ActivityQuestion.objects.filter(activity=activity).aggregate(total_max_score=Sum('score'))['total_max_score'] or 0 for activity in activities]),
                            'percentage': round((sum(activity.total_score for activity in student_activities.filter(activity__in=activities)) / sum(max_score for max_score in [ActivityQuestion.objects.filter(activity=activity).aggregate(total_max_score=Sum('score'))['total_max_score'] or 0 for activity in activities])) * 100, 2) if sum(max_score for max_score in [ActivityQuestion.objects.filter(activity=activity).aggregate(total_max_score=Sum('score'))['total_max_score'] or 0 for activity in activities]) > 0 else 0,
                        } for activity_type, activities in Activity.objects.filter(term=term, subject=subject, status=True).values_list('activity_type').distinct()
                    },
                    'attendance': {
                        'records': [
                            {
                                'date': record.date,
                                'status': record.status,
                            } for record in Attendance.objects.filter(student=student, term=term, subject=subject)
                        ],
                        'total_score': Attendance.objects.filter(student=student, term=term, subject=subject).aggregate(total_score=Sum('score'))['total_score'] or 0,
                        'max_score': Attendance.objects.filter(student=student, term=term, subject=subject).aggregate(total_max_score=Sum('max_score'))['total_max_score'] or 0,
                        'percentage': round((Attendance.objects.filter(student=student, term=term, subject=subject).aggregate(total_score=Sum('score'))['total_score'] or 0 / Attendance.objects.filter(student=student, term=term, subject=subject).aggregate(total_max_score=Sum('max_score'))['total_max_score'] or 0) * 100, 2) if Attendance.objects.filter(student=student, term=term, subject=subject).aggregate(total_max_score=Sum('max_score'))['total_max_score'] or 0 > 0 else 0,
                    },
                    'total_score': term_total_score,
                    'max_score': term_max_score,
                    'percentage': round((term_total_score / term_max_score) * 100, 2) if term_max_score > 0 else 0,
                } for term in terms
            }
        }
    }


@login_required
def grades(request):
    """Teacher-facing summary: pick a subject, see every enrolled student's
    per-term grade and final grade. Numbers come from the same helper that
    backs `/api/student-assessment-summary/` so values match the API contract.
    """
    from course.models import SubjectEnrollment, Term, Attendance
    from activity.utils.grade_calculation_utils import get_student_activity_summary

    semesters = Semester.objects.all().order_by('-start_date')
    current_semester = Semester.objects.filter(
        start_date__lte=now(), end_date__gte=now()
    ).first()

    selected_semester_id = request.GET.get('semester')
    if selected_semester_id:
        try:
            selected_semester = Semester.objects.filter(pk=int(selected_semester_id)).first() or current_semester
        except (TypeError, ValueError):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    user_role = request.user.role_name
    is_teacher = request.user.is_teacher

    # COIL/HALI courses run on a separate program-level grading flow and
    # never appear in the regular Grades view — for any role.
    subjects_qs = (
        Subject.objects.filter(subjectenrollment__semester=selected_semester)
        .exclude(is_coil=True).exclude(is_hali=True)
        .distinct()
    ) if selected_semester else Subject.objects.none()
    if is_teacher:
        subjects_qs = subjects_qs.filter(assign_teacher=request.user)
    subjects = list(subjects_qs.order_by('subject_name'))

    selected_subject = None
    selected_subject_id = request.GET.get('subject')
    if selected_subject_id:
        try:
            selected_subject = next((s for s in subjects if s.id == int(selected_subject_id)), None)
        except (TypeError, ValueError):
            selected_subject = None

    term_columns = []
    term_meta = []
    rows = []
    passing_grade = float(selected_semester.passing_grade) if selected_semester else 75.0

    if selected_semester and selected_subject:
        terms = list(Term.objects.filter(semester=selected_semester).order_by('start_date'))
        term_columns = [t.term_name for t in terms]
        visibility_settings = {
            v.term_id: v.is_visible
            for v in GradeVisibilitySettings.objects.filter(
                teacher=request.user,
                subject=selected_subject,
                term__in=terms,
            )
        }
        term_meta = [
            {
                'id': t.id,
                'name': t.term_name,
                'is_visible': visibility_settings.get(t.id, False),
            }
            for t in terms
        ]

        student_activities_qs = StudentActivity.objects.select_related(
            'activity', 'activity__activity_type', 'activity__subject', 'term', 'student__profile'
        ).filter(
            term__semester=selected_semester,
            activity__subject=selected_subject,
            activity__status=True,
        )
        attendance_qs = Attendance.objects.select_related('subject', 'student').filter(
            subject=selected_subject,
            graded=True,
            date__range=(selected_semester.start_date, selected_semester.end_date),
        )
        summary = get_student_activity_summary(
            selected_semester, selected_subject, terms,
            student_activities_qs, attendance_qs, request.user,
        )

        enrollments = SubjectEnrollment.objects.filter(
            subject=selected_subject, semester=selected_semester, status='enrolled'
        ).select_related('student__profile').order_by('student__profile__last_name', 'student__profile__first_name')

        for enr in enrollments:
            student = enr.student
            profile = getattr(student, 'profile', None)
            if not profile:
                continue
            student_name = f"{profile.first_name} {profile.last_name}"
            data = summary.get(student_name) or {}
            term_grades_map = data.get('term_grades') or {}

            term_breakdown = []
            for term_name in term_columns:
                term_data = term_grades_map.get(term_name) or {}
                term_breakdown.append({
                    'name': term_name,
                    'total': term_data.get('total_grade'),
                })

            final_grade = data.get('final_grade')
            try:
                final_grade_f = float(final_grade) if final_grade is not None else None
            except (TypeError, ValueError):
                final_grade_f = None

            rows.append({
                'student': student,
                'profile': profile,
                'terms': term_breakdown,
                'final_grade': final_grade_f,
                'is_passing': (final_grade_f is not None and final_grade_f >= passing_grade),
            })

    return render(request, 'grades/grades.html', {
        'semesters': semesters,
        'selected_semester': selected_semester,
        'subjects': subjects,
        'selected_subject': selected_subject,
        'rows': rows,
        'term_columns': term_columns,
        'term_meta': term_meta,
        'passing_grade': passing_grade,
    })


@login_required
def get_used_activity_types(request):
    subject_id = request.GET.get('subject_id')
    term_id = request.GET.get('term_id')
    exclude_gradebook = request.GET.get('exclude_gradebook')

    if not subject_id or not term_id:
        return JsonResponse({'used_ids': []})

    qs = ActivityTypePercentage.objects.filter(
        gradebook_component__teacher=request.user,
        gradebook_component__subject_id=subject_id,
        gradebook_component__term_id=term_id,
    )
    if exclude_gradebook:
        qs = qs.exclude(gradebook_component_id=exclude_gradebook)

    used_ids = list(qs.values_list('activity_type_id', flat=True).distinct())
    return JsonResponse({'used_ids': used_ids})
