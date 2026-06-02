from django.shortcuts import render, redirect, get_object_or_404
from course.models import SubjectEnrollment, Semester, Attendance, AttendanceStatus, TeacherAttendancePoints
from accounts.models import Course
from subject.models import Subject
from accounts.models import CustomUser
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import permission_required
from datetime import datetime, timedelta
from django.http import JsonResponse
from django.utils.dateparse import parse_date
from django.db import IntegrityError
from django.db.models import Count
import calendar 
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from course.forms import AttendanceForm, TeacherAttendancePointsForm, updateAttendanceForm
from activity.utils.authorization import check_subject_access
from subject.models import Schedule
from django.core.paginator import Paginator
from django.db.models import Q

@login_required
@permission_required('course.add_attendance', raise_exception=True)
def record_attendanceCM(request, subject_id):
    current_date = timezone.now().date()
    current_semester = Semester.objects.filter(
        start_date__lte=current_date,
        end_date__gte=current_date,
        end_semester=False
    ).first()

    schedule = Schedule.objects.filter(semester=current_semester).first()

    if not current_semester:
        return redirect('404.html')

    enrollments = SubjectEnrollment.objects.filter(
        subject_id=subject_id,
        semester=current_semester,
        status='enrolled'
    ).select_related('student').order_by('student__last_name', 'student__first_name')
    
    students = [enrollment.student for enrollment in enrollments]
    attendance_statuses = AttendanceStatus.objects.all()
    subject = get_object_or_404(Subject, id=subject_id)
    teacher = request.user  
    
    # Authorization: only assigned teachers/collaborators/admins may record attendance
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    
    # Add points to each status for the current teacher
    for status in attendance_statuses:
        teacher_points = TeacherAttendancePoints.objects.filter(teacher=teacher, status=status).first()
        status.points = teacher_points.points if teacher_points else 0

    if request.method == 'POST':
        selected_date = request.POST.get('date')
        graded_attendance = request.POST.get('graded') == 'on'  

        if not selected_date:
            messages.error(request, 'Date is required to record attendance.')
            return redirect('record_attendanceCM', subject_id=subject_id)

        missing_status_students = []

        existing_attendance = Attendance.objects.filter(subject_id=subject_id, date=selected_date)
        attendance_by_student = {att.student_id: att for att in existing_attendance}
        already_recorded_students = []
        new_records = 0

        for user_id in request.POST.getlist('students'):
            # Skip empty or invalid user IDs
            if not user_id or not str(user_id).strip():
                continue
            
            try:
                user_id = int(user_id)
                student = CustomUser.objects.get(id=user_id)
            except (ValueError, CustomUser.DoesNotExist):
                continue

            status_value = request.POST.get(f'status_{user_id}')
            remark_value = request.POST.get(f'remark_{user_id}', '')

            if student.id in attendance_by_student:
                already_recorded_students.append(student)
                continue

            if status_value and status_value.strip():
                try:
                    status = AttendanceStatus.objects.get(id=status_value)
                except AttendanceStatus.DoesNotExist:
                    return redirect('record_attendanceCM', subject_id=subject_id)

                attendance, created = Attendance.objects.update_or_create(
                    student_id=student.id,
                    subject_id=subject_id,
                    date=selected_date,
                    defaults={
                        'status': status,
                        'remark': remark_value,
                        'graded': graded_attendance, 
                        'teacher': teacher,
                        'schedule': schedule,
                    }
                )
                if created:
                    new_records += 1
                attendance_by_student[student.id] = attendance
            else:
                missing_status_students.append(student)

        # If any students are missing status, show an error and prevent submission
        if missing_status_students:
            messages.error(request, 'Some students have missing attendance statuses.')
            return redirect('classroom_mode', subject_id=subject_id)

        if new_records == 0:
            messages.info(request, 'All selected students already have attendance recorded for this date.')
        else:
            messages.success(request, 'Attendance recorded successfully!')
        return redirect('classroom_mode', pk=subject_id)

    else:
        form = AttendanceForm(current_semester=current_semester, subject=subject_id)

    context = {
        'students': students,  
        'attendance_statuses': attendance_statuses,
        'subject': subject,
        'form': form,
    }

    return render(request, 'course/attendance/teacherAttendanceCM.html', context)

@login_required
@permission_required('course.add_attendance', raise_exception=True)
def record_attendance(request, subject_id):
    current_date = timezone.now().date()
    current_semester = Semester.objects.filter(
        start_date__lte=current_date,
        end_date__gte=current_date,
        end_semester=False
    ).first()
    
    schedule = Schedule.objects.filter(semester=current_semester).first()

    if not current_semester:
        return redirect('404.html')

    enrollments = SubjectEnrollment.objects.filter(
        subject_id=subject_id,
        semester=current_semester,
        status='enrolled'
    ).select_related('student').order_by('student__last_name', 'student__first_name')  
    students = [enrollment.student for enrollment in enrollments]
    attendance_statuses = AttendanceStatus.objects.all()
    subject = get_object_or_404(Subject, id=subject_id)
    teacher = request.user  # Set the current teacher
    
    # Authorization: only assigned teachers/collaborators/admins may record attendance
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    
    # Add points to each status for the current teacher
    for status in attendance_statuses:
        teacher_points = TeacherAttendancePoints.objects.filter(teacher=teacher, status=status).first()
        status.points = teacher_points.points if teacher_points else 0

    if request.method == 'POST':
        selected_date = request.POST.get('date')
        graded_attendance = request.POST.get('graded') == 'on'  # Check if attendance should be graded

        if not selected_date:
            messages.error(request, 'Date is required to record attendance.')
            return redirect('record-attendance', subject_id=subject_id)

        parsed_date = parse_date(selected_date)
        if not parsed_date:
            messages.error(request, 'Invalid date format. Please pick a valid date.')
            return redirect('record-attendance', subject_id=subject_id)
        if parsed_date > current_date:
            messages.error(request, 'Attendance date cannot be in the future.')
            return redirect('record-attendance', subject_id=subject_id)
        if (parsed_date < current_semester.start_date or parsed_date > current_semester.end_date):
            messages.error(request, 'Attendance date must fall within the current semester.')
            return redirect('record-attendance', subject_id=subject_id)

        missing_status_students = []

        existing_attendance = Attendance.objects.filter(subject_id=subject_id, date=selected_date)
        attendance_by_student = {att.student_id: att for att in existing_attendance}
        already_recorded_students = []
        new_records = 0

        for user_id in request.POST.getlist('students'):
            # Skip empty or invalid user IDs
            if not user_id or not str(user_id).strip():
                continue
            
            try:
                user_id = int(user_id)
                student = CustomUser.objects.get(id=user_id)
            except (ValueError, CustomUser.DoesNotExist):
                continue

            status_value = request.POST.get(f'status_{user_id}')
            remark_value = request.POST.get(f'remark_{user_id}', '')

            if student.id in attendance_by_student:
                already_recorded_students.append(student)
                continue

            if status_value and status_value.strip():
                try:
                    status = AttendanceStatus.objects.get(id=status_value)
                except AttendanceStatus.DoesNotExist:
                    return redirect('attendance-report')

                # Create or update attendance, set teacher and graded fields
                attendance, created = Attendance.objects.update_or_create(
                    student_id=student.id,
                    subject_id=subject_id,
                    date=selected_date,
                    defaults={
                        'status': status,
                        'remark': remark_value,
                        'graded': graded_attendance,
                        'teacher': teacher,
                        'schedule': schedule,
                    }
                )
                if created:
                    new_records += 1
                attendance_by_student[student.id] = attendance
            else:
                missing_status_students.append(student)

        # If any students are missing status, show an error and prevent submission
        if missing_status_students:
            messages.error(request, 'Some students have missing attendance statuses.')
            return redirect('attendance-report')
        
        if new_records == 0:
            messages.info(request, 'All selected students already have attendance recorded for this date.')
        else:
            messages.success(request, 'Attendance recorded successfully!')
        return redirect('attendance-report')

    else:
        form = AttendanceForm(current_semester=current_semester, subject=subject_id)

    context = {
        'students': students,  
        'attendance_statuses': attendance_statuses,
        'subject': subject,
        'form': form,
    }

    return render(request, 'course/attendance/record-attendance.html', context)


@login_required
@permission_required('course.view_attendance', raise_exception=True)
def attendance_list(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    current_date = timezone.now().date()

    selected_date_str = request.GET.get('date', None)
    if selected_date_str:
        # Ensure the date is parsed only if a valid string is passed
        selected_date = parse_date(selected_date_str)
        if not selected_date:
            messages.error(request, 'Invalid date format. Please select a valid date.')
            selected_date = current_date
    else:
        selected_date = current_date

    attendance = Attendance.objects.filter(subject_id=subject_id, date=selected_date).order_by('-date')
    available_dates = Attendance.objects.filter(subject_id=subject_id).values_list('date', flat=True).distinct()

    return render(request, 'course/attendance/attendance_list.html', {
        'attendance': attendance,
        'selected_date': selected_date,
        'available_dates': available_dates,
    })

@login_required
@permission_required('course.view_attendance', raise_exception=True)
def attendance_list_CM(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    current_date = timezone.now().date()

    selected_date_str = request.GET.get('date', None)
    if selected_date_str:
        # Ensure the date is parsed only if a valid string is passed
        selected_date = parse_date(selected_date_str)
        if not selected_date:
            messages.error(request, 'Invalid date format. Please select a valid date.')
            selected_date = current_date
    else:
        selected_date = current_date

    attendance = Attendance.objects.filter(subject_id=subject_id, date=selected_date).order_by('-date')
    available_dates = Attendance.objects.filter(subject_id=subject_id).values_list('date', flat=True).distinct()

    return render(request, 'course/attendance/attendance_list_CM.html', {
        'attendance': attendance,
        'subject': subject,
        'selected_date': selected_date,
        'available_dates': available_dates,
    })


@login_required
@permission_required('course.change_attendance', raise_exception=True)
def update_attendance(request, id):
    attendance = get_object_or_404(Attendance, id=id)
    subject = attendance.subject
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    form = updateAttendanceForm(instance=attendance)
    teacher = request.user  # Set the teacher as the current user

    if request.method == 'POST':
        form = updateAttendanceForm(request.POST, instance=attendance)
        graded_attendance = request.POST.get('graded') == 'on'  # Check if graded checkbox is checked

        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.teacher = teacher  # Update teacher
            attendance.graded = graded_attendance  # Update the graded status
            attendance.save()  # Save changes
            messages.success(request, 'Attendance updated successfully!')
            return redirect('attendanceList', subject_id=attendance.subject.id)
        else:
            messages.error(request, 'There was an error updating the attendance. Please try again.')
            return redirect('attendanceList', subject_id=attendance.subject.id)

    return render(request, 'course/attendance/update-attendance.html', { 
        'form': form, 
        'attendance': attendance 
    })

@login_required
@permission_required('course.change_attendance', raise_exception=True)
def update_attendance_CM(request, id):
    attendance = get_object_or_404(Attendance, id=id)
    subject = attendance.subject
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    form = updateAttendanceForm(instance=attendance)
    teacher = request.user  # Set the teacher as the current user

    if request.method == 'POST':
        form = updateAttendanceForm(request.POST, instance=attendance)
        graded_attendance = request.POST.get('graded') == 'on'  # Check if graded checkbox is checked

        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.teacher = teacher  # Update teacher
            attendance.graded = graded_attendance  # Update the graded status
            attendance.save()  # Save changes
            messages.success(request, 'Attendance updated successfully!')
            return redirect('attendanceListCM', subject_id=attendance.subject.id)
        else:
            messages.error(request, 'There was an error updating the attendance. Please try again.')
            return redirect('attendanceListCM', subject_id=attendance.subject.id)

    return render(request, 'course/attendance/update_attendance_CM.html', { 
        'form': form, 
        'attendance': attendance 
    })



@login_required
def status_list(request):
    form = TeacherAttendancePointsForm()
    user = request.user
    status_points = TeacherAttendancePoints.objects.filter(teacher=user)
    return render(request, 'course/status/status-list.html', {'status_points': status_points, 'form': form})


@login_required
def add_points(request):
    is_xhr = request.headers.get('x-requested-with') == 'XMLHttpRequest'

    if request.method == 'POST':
        form = TeacherAttendancePointsForm(request.POST)
        if form.is_valid():
            status = form.cleaned_data['status']
            points = form.cleaned_data['points']
            teacher = request.user

            try:
                instance, created = TeacherAttendancePoints.objects.update_or_create(
                    teacher=teacher,
                    status=status,
                    defaults={'points': points}
                )
                if is_xhr:
                    return JsonResponse({
                        'success': True,
                        'message': 'Status points assigned successfully!' if created else 'Status points updated successfully!'
                    })
                messages.success(request, 'Status points assigned successfully!' if created else 'Status points updated successfully!')
                return redirect('status-list')
            except IntegrityError:
                if is_xhr:
                    return JsonResponse({'success': False, 'message': 'An error occurred while saving points. Please try again.'}, status=400)
                messages.error(request, 'An error occurred while saving points. Please try again.')

        else:
            if is_xhr:
                return JsonResponse({'success': False, 'errors': form.errors}, status=400)
    else:
        form = TeacherAttendancePointsForm()

    return render(request, 'course/status/add-points.html', {'form': form})


@login_required
def update_points(request, id):
    status_points = get_object_or_404(TeacherAttendancePoints, id=id, teacher=request.user)

    if request.method == 'POST':
        form = TeacherAttendancePointsForm(request.POST, instance=status_points)
        if form.is_valid():
            form.save()
            messages.success(request, 'Status points updated successfully!')
            return redirect('status-list')
    else:
        form = TeacherAttendancePointsForm(instance=status_points)

    return render(request, 'course/status/update-points.html', {'form': form, 'status_points': status_points})


@login_required
def delete_points(request, id):
    status_points = get_object_or_404(TeacherAttendancePoints, id=id, teacher=request.user)

    if request.method == 'POST':
        status_points.delete()
        messages.success(request, 'Status points deleted successfully!')
        return redirect('status-list')

    return render(request, 'course/status/delete_points.html', {'status_points': status_points})


@login_required
def get_student_attendance(course_id=None, subject_id=None):
    """
    Fetch student attendance records based on course and/or subject.
    
    :param course_id: ID of the course (Optional)
    :param subject_id: ID of the subject (Optional)
    :return: Queryset of attendance records
    """

    # Start with all attendance records
    attendance_records = Attendance.objects.all()

    # Filter by course if provided
    if course_id:
        attendance_records = attendance_records.filter(student__profile__course_id=course_id)

    # Filter by subject if provided
    if subject_id:
        attendance_records = attendance_records.filter(subject_id=subject_id)

    # Optimize query by preloading related fields
    attendance_records = attendance_records.select_related(
        'student', 'subject', 'status', 'student__profile'
    )

    return attendance_records


@login_required
def attendance_report(request):

    course_id = request.GET.get('course_id')
    subject_id = request.GET.get('subject_id')
    status_id = request.GET.get('status_id')
    from_date_str = request.GET.get('from_date')
    to_date_str = request.GET.get('to_date')
    search_query = (request.GET.get('q') or '').strip()
    try:
        per_page = int(request.GET.get('per_page', 25))
    except (TypeError, ValueError):
        per_page = 25
    if per_page not in (10, 25, 50, 100):
        per_page = 25

    # Default the date range to today on initial visit (no GET params).
    # Once the form is submitted, empty date fields mean "any date".
    if not request.GET:
        today_str = timezone.localdate().isoformat()
        from_date_str = today_str
        to_date_str = today_str

    from_date = parse_date(from_date_str) if from_date_str else None
    to_date = parse_date(to_date_str) if to_date_str else None

    courses = Course.objects.all()
    if getattr(request.user, 'role_name', None) == 'teacher':
        subjects = Subject.objects.filter(
            Q(assign_teacher=request.user)
            | Q(allow_substitute_teacher=True, substitute_teacher=request.user)
        ).distinct()
    else:
        subjects = Subject.objects.all()
    attendance_status_list = AttendanceStatus.objects.all()

    attendance_records = (
        Attendance.objects.all()
        .select_related('student', 'subject', 'status', 'student__profile')
        .order_by('-date', 'student__last_name', 'student__first_name')
    )

    if course_id:
        attendance_records = attendance_records.filter(student__profile__course_id=course_id)
    if subject_id:
        attendance_records = attendance_records.filter(subject_id=subject_id)
    if status_id:
        attendance_records = attendance_records.filter(status_id=status_id)
    if from_date:
        attendance_records = attendance_records.filter(date__gte=from_date)
    if to_date:
        attendance_records = attendance_records.filter(date__lte=to_date)
    if search_query:
        attendance_records = attendance_records.filter(
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(student__username__icontains=search_query) |
            Q(subject__subject_name__icontains=search_query) |
            Q(status__status__icontains=search_query)
        )

    total_count = attendance_records.count()

    # Stats by status bucket for the cards row. Bucket labels are lowercase
    # substrings so they match variants like "Present_Online" and "Tardy".
    status_breakdown = list(
        attendance_records.values('status__status').annotate(total=Count('id')).order_by('-total')
    )
    bucket_counts = {'present': 0, 'absent': 0, 'late': 0, 'excused': 0, 'other': 0}
    for row in status_breakdown:
        label = (row['status__status'] or '').lower()
        n = row['total']
        if 'present' in label:
            bucket_counts['present'] += n
        elif 'absent' in label:
            bucket_counts['absent'] += n
        elif 'late' in label or 'tardy' in label:
            bucket_counts['late'] += n
        elif 'excus' in label:
            bucket_counts['excused'] += n
        else:
            bucket_counts['other'] += n

    def _pct(n):
        return round((n / total_count) * 100, 1) if total_count else 0

    stats = {
        'total': total_count,
        'present': bucket_counts['present'],
        'absent': bucket_counts['absent'],
        'late': bucket_counts['late'],
        'excused': bucket_counts['excused'],
        'other': bucket_counts['other'],
        'present_pct': _pct(bucket_counts['present']),
        'absent_pct': _pct(bucket_counts['absent']),
        'late_pct': _pct(bucket_counts['late']),
        'excused_pct': _pct(bucket_counts['excused']),
        'other_pct': _pct(bucket_counts['other']),
        'unique_students': attendance_records.values('student_id').distinct().count(),
        'unique_subjects': attendance_records.values('subject_id').distinct().count(),
    }

    paginator = Paginator(attendance_records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    qs_params = request.GET.copy()
    qs_params.pop('page', None)
    querystring = qs_params.urlencode()

    active_filter_count = sum(1 for v in (course_id, subject_id, status_id, from_date_str, to_date_str) if v)

    # Today's classes strip for the teacher — drives the self-attendance
    # open/close toggle inline in the report (same UX as record-attendance).
    ENABLE_LEAD_MINUTES = 15
    now = timezone.localtime()
    today = now.date()
    today_abbr = now.strftime("%a")
    today_classes = []
    teacher_subjects = Subject.objects.filter(
        Q(assign_teacher=request.user)
        | Q(allow_substitute_teacher=True, substitute_teacher=request.user)
    ).distinct()
    for subj in teacher_subjects:
        for sched in subj.schedules.filter(is_active_semester=True):
            if today_abbr not in sched.days_of_week:
                continue
            start_dt = datetime.combine(today, sched.schedule_start_time)
            end_dt = datetime.combine(today, sched.schedule_end_time)
            open_at = start_dt - timedelta(minutes=ENABLE_LEAD_MINUTES)
            now_naive = now.replace(tzinfo=None)
            today_classes.append({
                'subject': subj,
                'schedule_start': sched.schedule_start_time,
                'schedule_end': sched.schedule_end_time,
                'self_on': subj.self_attendance_enabled,
                'window_open': open_at <= now_naive <= end_dt,
            })
    today_classes.sort(key=lambda c: c['schedule_start'])

    return render(request, 'course/attendance/attendance-report.html', {
        'today_classes': today_classes,
        'attendance_records': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_count': total_count,
        'courses': courses,
        'subjects': subjects,
        'attendance_status_list': attendance_status_list,
        'search_query': search_query,
        'per_page': per_page,
        'querystring': querystring,
        'stats': stats,
        'from_date': from_date_str or '',
        'to_date': to_date_str or '',
        'active_filter_count': active_filter_count,
    })



@login_required
@permission_required('course.view_attendance', raise_exception=True)
def student_subject_attendance(request, subject_id, student_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    student = get_object_or_404(CustomUser, pk=student_id)
    
    # Students can only view their own attendance
    user_role = request.user.role_name
    if user_role == 'student' and request.user.id != student.id:
        messages.error(request, "You can only view your own attendance.")
        return redirect('subjectStudentList', pk=subject_id)
    
    # Check subject access authorization
    has_access, redirect_response = check_subject_access(request, subject)
    if not has_access:
        return redirect_response
    
    # Verify student is enrolled in this subject (prevents teachers from viewing non-enrolled students)
    if user_role in ['teacher', 'student']:
        now = timezone.localtime(timezone.now())
        current_semester = Semester.objects.filter(start_date__lte=now, end_date__gte=now).first()
        is_enrolled = SubjectEnrollment.objects.filter(
            student=student,
            subject=subject,
            semester=current_semester
        ).exists()
        
        if not is_enrolled:
            messages.error(request, "This student is not enrolled in this subject.")
            return redirect('subjectStudentList', pk=subject_id)

    # Semester filter (optional)
    selected_semester = None
    semester_id = request.GET.get('semester')
    if semester_id:
        selected_semester = get_object_or_404(Semester, pk=semester_id)
        base_qs = Attendance.objects.filter(
            subject=subject,
            student=student,
            date__range=(selected_semester.start_date, selected_semester.end_date),
        )
    else:
        base_qs = Attendance.objects.filter(subject=subject, student=student)

    # Attendance doesn't have a direct `schedule` FK; selecting it triggers errors in CM view.
    base_qs = base_qs.select_related('status', 'teacher').order_by('date')

    # Summary counts (for both views)
    status_counts = (
        base_qs.values('status__status')
        .annotate(total=Count('id'))
        .order_by('status__status')
    )

    # Build points map (teacher,status) -> points
    teacher_ids = base_qs.values_list('teacher_id', flat=True).distinct()
    tap_map = {}
    for tap in TeacherAttendancePoints.objects.filter(teacher_id__in=teacher_ids):
        tap_map[(tap.teacher_id, tap.status_id)] = tap.points

    # Table rows list of tuples (rec, pts)
    attendance_rows = []
    total_points = 0
    for rec in base_qs:
        pts = tap_map.get((rec.teacher_id, rec.status_id), 0)
        attendance_rows.append((rec, pts))
        if rec.graded:
            total_points += pts

    # -------- Calendar Mode Prep --------
    view_mode = request.GET.get('view', 'table')  # 'table' or 'calendar'

    # Which month?
    month_param = request.GET.get('month')
    if month_param:
        try:
            year, month = map(int, month_param.split('-', 1))
        except Exception:
            today = timezone.localdate()
            year, month = today.year, today.month
    else:
        today = timezone.localdate()
        year, month = today.year, today.month

    # (Optional) clamp to semester range
    if selected_semester:
        if year < selected_semester.start_date.year or \
           (year == selected_semester.start_date.year and month < selected_semester.start_date.month) or \
           year > selected_semester.end_date.year or \
           (year == selected_semester.end_date.year and month > selected_semester.end_date.month):
            year = selected_semester.start_date.year
            month = selected_semester.start_date.month

    # ----- Status visuals -----
    # If AttendanceStatus has a 'color' and/or 'abbr' field, pull from DB here.
    # Otherwise, define fallback maps:
    STATUS_COLORS = {
        'Present': '#00A859',
        'Present_Online': '#00A859',
        'Absent': '#AF1E2D',
        'Late': '#FFC72C',
        'Excused': '#00B5E2',
        
    }
    STATUS_ABBR = {
        'Present': 'P',
        'Present_Online': 'PO',
        'Absent': 'A',
        'Late': 'L',
        'Excused': 'E',
        
    }

    # Build dict date->attendance record
    attendance_by_date = {rec.date: rec for rec in base_qs}

    # Build weeks grid (Sunday start; 6 = Sunday; change to 0 for Monday)
    cal = calendar.Calendar(firstweekday=6)
    weeks = []
    for wk in cal.monthdatescalendar(year, month):
        row = []
        for d in wk:
            rec = attendance_by_date.get(d)
            if rec and rec.status:
                label = rec.status.status
                color = STATUS_COLORS.get(label, '#6C757D')
                short = STATUS_ABBR.get(label, label[:1].upper())
            else:
                label = ''
                color = None
                short = ''
            row.append({
                'date': d,
                'in_month': (d.month == month),
                'rec': rec,
                'color': color,
                'label': label,
                'short': short,
            })
        weeks.append(row)

    month_name = calendar.month_name[month]

    context = {
        'subject': subject,
        'student': student,
        'attendance_rows': attendance_rows,
        'status_counts': status_counts,
        'selected_semester': selected_semester,
        'total_points': total_points,
        'view_mode': view_mode,
        'year': year,
        'month': month,
        'weeks': weeks,
        'status_colors': STATUS_COLORS,  # used for legend
        'month_name': month_name,
    }
    return render(request, 'course/attendance/student_subject_attendance.html', context)


@login_required
@permission_required('course.view_attendance', raise_exception=True)
def student_subject_attendanceCM(request, subject_id, student_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    student = get_object_or_404(CustomUser, pk=student_id)
    
    # Students can only view their own attendance
    user_role = request.user.role_name
    if user_role == 'student' and request.user.id != student.id:
        messages.error(request, "You can only view your own attendance.")
        return redirect('subjectStudentListCM', pk=subject_id)
    
    # Check subject access authorization
    has_access, redirect_response = check_subject_access(request, subject)
    if not has_access:
        return redirect_response
    
    # Verify student is enrolled in this subject (prevents teachers from viewing non-enrolled students)
    if user_role in ['teacher', 'student']:
        now = timezone.localtime(timezone.now())
        current_semester = Semester.objects.filter(start_date__lte=now, end_date__gte=now).first()
        is_enrolled = SubjectEnrollment.objects.filter(
            student=student,
            subject=subject,
            semester=current_semester
        ).exists()
        
        if not is_enrolled:
            messages.error(request, "This student is not enrolled in this subject.")
            return redirect('subjectStudentListCM', pk=subject_id)

    # Semester filter (optional)
    selected_semester = None
    semester_id = request.GET.get('semester')
    if semester_id:
        selected_semester = get_object_or_404(Semester, pk=semester_id)
        base_qs = Attendance.objects.filter(
            subject=subject,
            student=student,
            date__range=(selected_semester.start_date, selected_semester.end_date),
        )
    else:
        base_qs = Attendance.objects.filter(subject=subject, student=student)

    base_qs = base_qs.select_related('status', 'teacher').order_by('date')

    # Summary counts (for both views)
    status_counts = (
        base_qs.values('status__status')
        .annotate(total=Count('id'))
        .order_by('status__status')
    )

    # Build points map (teacher,status) -> points
    teacher_ids = base_qs.values_list('teacher_id', flat=True).distinct()
    tap_map = {}
    for tap in TeacherAttendancePoints.objects.filter(teacher_id__in=teacher_ids):
        tap_map[(tap.teacher_id, tap.status_id)] = tap.points

    # Table rows list of tuples (rec, pts)
    attendance_rows = []
    total_points = 0
    for rec in base_qs:
        pts = tap_map.get((rec.teacher_id, rec.status_id), 0)
        attendance_rows.append((rec, pts))
        if rec.graded:
            total_points += pts

    # -------- Calendar Mode Prep --------
    view_mode = request.GET.get('view', 'table')  # 'table' or 'calendar'

    # Which month?
    month_param = request.GET.get('month')
    if month_param:
        try:
            year, month = map(int, month_param.split('-', 1))
        except Exception:
            today = timezone.localdate()
            year, month = today.year, today.month
    else:
        today = timezone.localdate()
        year, month = today.year, today.month

    # (Optional) clamp to semester range
    if selected_semester:
        if year < selected_semester.start_date.year or \
           (year == selected_semester.start_date.year and month < selected_semester.start_date.month) or \
           year > selected_semester.end_date.year or \
           (year == selected_semester.end_date.year and month > selected_semester.end_date.month):
            year = selected_semester.start_date.year
            month = selected_semester.start_date.month

    # ----- Status visuals -----
    # If AttendanceStatus has a 'color' and/or 'abbr' field, pull from DB here.
    # Otherwise, define fallback maps:
    STATUS_COLORS = {
        'Present': '#00A859',
        'Present_Online': '#00A859',
        'Absent': '#AF1E2D',
        'Late': '#FFC72C',
        'Excused': '#00B5E2',
        
    }
    STATUS_ABBR = {
        'Present': 'P',
        'Present_Online': 'PO',
        'Absent': 'A',
        'Late': 'L',
        'Excused': 'E',
        
    }

    # Build dict date->attendance record
    attendance_by_date = {rec.date: rec for rec in base_qs}

    # Build weeks grid (Sunday start; 6 = Sunday; change to 0 for Monday)
    cal = calendar.Calendar(firstweekday=6)
    weeks = []
    for wk in cal.monthdatescalendar(year, month):
        row = []
        for d in wk:
            rec = attendance_by_date.get(d)
            if rec and rec.status:
                label = rec.status.status
                color = STATUS_COLORS.get(label, '#6C757D')
                short = STATUS_ABBR.get(label, label[:1].upper())
            else:
                label = ''
                color = None
                short = ''
            row.append({
                'date': d,
                'in_month': (d.month == month),
                'rec': rec,
                'color': color,
                'label': label,
                'short': short,
            })
        weeks.append(row)

    month_name = calendar.month_name[month]

    context = {
        'subject': subject,
        'student': student,
        'attendance_rows': attendance_rows,
        'status_counts': status_counts,
        'selected_semester': selected_semester,
        'total_points': total_points,
        'view_mode': view_mode,
        'year': year,
        'month': month,
        'weeks': weeks,
        'status_colors': STATUS_COLORS,  # used for legend
        'month_name': month_name,
    }
    return render(request, 'course/attendance/student_subject_attendanceCM.html', context)


@login_required
def export_attendance_by_date_range(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    # Role check
    role_name = request.user.role_name
    is_teacher = role_name == 'teacher'
    is_staff = request.user.is_staff

    if not (is_teacher or is_staff):
        messages.error(request, "You don't have permission to export attendance data.")
        return redirect('subjectStudentList', pk=subject_id)

    # Date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        end_date = timezone.localdate()
        start_date = end_date - timedelta(days=6)

    # Semester
    selected_semester_id = request.GET.get('semester')
    if selected_semester_id and selected_semester_id != 'None':
        selected_semester = get_object_or_404(Semester, id=selected_semester_id)
    else:
        selected_semester = Semester.objects.filter(
            start_date__lte=timezone.now().date(),
            end_date__gte=timezone.now().date()
        ).first()

    if not selected_semester:
        messages.error(request, "No active semester found.")
        return redirect('subjectStudentList', pk=subject_id)

    students = CustomUser.objects.filter(
        subjectenrollment__subject=subject,
        subjectenrollment__semester=selected_semester,
        subjectenrollment__status='enrolled'
    ).distinct().order_by('last_name', 'first_name')

    # Generate date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Add title and information rows
    ws.append([f"Attendance Report: {subject.subject_name}"])
    ws.append([f"Semester: {selected_semester.semester_name} - {selected_semester.start_date.year if selected_semester.start_date.year == selected_semester.end_date.year else f'{selected_semester.start_date.year}-{selected_semester.end_date.year}'}"])
    ws.append([f"Date Range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"])
    ws.append([]) # Empty row for spacing
    
    # Make title rows bold
    for row in range(1, 4):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Header row
    header = ['Student Name'] + [date.strftime('%m/%d/%Y') for date in date_range]
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=5, column=col).font = Font(bold=True)

    # Data rows
    for student in students:
        row = [f"{student.last_name}, {student.first_name}"]

        attendance_records = Attendance.objects.filter(
            student=student,
            subject=subject,
            date__range=(start_date, end_date)
        ).select_related('status')

        attendance_by_date = {record.date: record.status.status if record.status else '--'
                              for record in attendance_records}

        for date in date_range:
            row.append(attendance_by_date.get(date, '--'))

        ws.append(row)

    # Add consolidated attendance statistics section
    ws.append([])  # Empty row for spacing
    ws.append([])  # Another empty row
    
    # Consolidated section title
    consolidated_row = ws.max_row + 1
    ws.append(["Consolidated Attendance Statistics"])
    ws.cell(row=consolidated_row, column=1).font = Font(bold=True, size=12)
    ws.append([])  # Empty row
    
    # Get all attendance statuses
    from course.models import AttendanceStatus
    all_statuses = AttendanceStatus.objects.all()
    
    # Consolidated header
    consolidated_header = ['Student Name'] + [status.status for status in all_statuses] + ['Total']
    consolidated_header_row = ws.max_row + 1
    ws.append(consolidated_header)
    for col in range(1, len(consolidated_header) + 1):
        ws.cell(row=consolidated_header_row, column=col).font = Font(bold=True)
    
    # Consolidated data rows
    for student in students:
        student_attendances = Attendance.objects.filter(
            student=student,
            subject=subject,
            date__range=(start_date, end_date)
        ).select_related('status')
        
        status_counts = {}
        total_count = 0
        
        for status in all_statuses:
            count = student_attendances.filter(status=status).count()
            status_counts[status.status] = count
            total_count += count
        
        consolidated_row = [f"{student.last_name}, {student.first_name}"]
        for status in all_statuses:
            consolidated_row.append(status_counts.get(status.status, 0))
        consolidated_row.append(total_count)
        
        ws.append(consolidated_row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"attendance_{subject.subject_name}_{selected_semester.semester_name}_{start_date}_to_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response


@login_required
def export_attendance_by_date_rangeCM(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    
    # ===== AUTHORIZATION CHECK =====
    has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
    if not has_access:
        return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    # Role check
    role_name = request.user.role_name
    is_teacher = role_name == 'teacher'
    is_staff = request.user.is_staff

    if not (is_teacher or is_staff):
        messages.error(request, "You don't have permission to export attendance data.")
        return redirect('subjectStudentListCM', pk=subject_id)

    # Date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        end_date = timezone.localdate()
        start_date = end_date - timedelta(days=6)

    # Semester
    selected_semester_id = request.GET.get('semester')
    if selected_semester_id and selected_semester_id != 'None':
        selected_semester = get_object_or_404(Semester, id=selected_semester_id)
    else:
        selected_semester = Semester.objects.filter(
            start_date__lte=timezone.now().date(),
            end_date__gte=timezone.now().date()
        ).first()

    if not selected_semester:
        messages.error(request, "No active semester found.")
        return redirect('subjectStudentListCM', pk=subject_id)

    students = CustomUser.objects.filter(
        subjectenrollment__subject=subject,
        subjectenrollment__semester=selected_semester,
        subjectenrollment__status='enrolled'
    ).distinct().order_by('last_name', 'first_name')

    # Generate date range
    date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

    # Create Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"
    
    # Add title and information rows
    ws.append([f"Attendance Report: {subject.subject_name}"])
    ws.append([f"Semester: {selected_semester.semester_name} - {selected_semester.start_date.year if selected_semester.start_date.year == selected_semester.end_date.year else f'{selected_semester.start_date.year}-{selected_semester.end_date.year}'}"])
    ws.append([f"Date Range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"])
    ws.append([]) # Empty row for spacing
    
    # Make title rows bold
    for row in range(1, 4):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Header row
    header = ['Student Name'] + [date.strftime('%m/%d/%Y') for date in date_range]
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=5, column=col).font = Font(bold=True)

    # Data rows
    for student in students:
        row = [f"{student.last_name}, {student.first_name}"]

        attendance_records = Attendance.objects.filter(
            student=student,
            subject=subject,
            date__range=(start_date, end_date)
        ).select_related('status')

        attendance_by_date = {record.date: record.status.status if record.status else '--'
                              for record in attendance_records}

        for date in date_range:
            row.append(attendance_by_date.get(date, '--'))

        ws.append(row)

    # Add consolidated attendance statistics section
    ws.append([])  # Empty row for spacing
    ws.append([])  # Another empty row
    
    # Consolidated section title
    consolidated_row = ws.max_row + 1
    ws.append(["Consolidated Attendance Statistics"])
    ws.cell(row=consolidated_row, column=1).font = Font(bold=True, size=12)
    ws.append([])  # Empty row
    
    # Get all attendance statuses
    from course.models import AttendanceStatus
    all_statuses = AttendanceStatus.objects.all()
    
    # Consolidated header
    consolidated_header = ['Student Name'] + [status.status for status in all_statuses] + ['Total']
    consolidated_header_row = ws.max_row + 1
    ws.append(consolidated_header)
    for col in range(1, len(consolidated_header) + 1):
        ws.cell(row=consolidated_header_row, column=col).font = Font(bold=True)
    
    # Consolidated data rows
    for student in students:
        student_attendances = Attendance.objects.filter(
            student=student,
            subject=subject,
            date__range=(start_date, end_date)
        ).select_related('status')
        
        status_counts = {}
        total_count = 0
        
        for status in all_statuses:
            count = student_attendances.filter(status=status).count()
            status_counts[status.status] = count
            total_count += count
        
        consolidated_row = [f"{student.last_name}, {student.first_name}"]
        for status in all_statuses:
            consolidated_row.append(status_counts.get(status.status, 0))
        consolidated_row.append(total_count)
        
        ws.append(consolidated_row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"attendance_{subject.subject_name}_{selected_semester.semester_name}_{start_date}_to_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response


@login_required
def export_student_attendance(request, subject_id, student_id):
    subject = get_object_or_404(Subject, id=subject_id)
    student = get_object_or_404(CustomUser, id=student_id)
    
    # ===== AUTHORIZATION CHECK =====
    user = request.user
    is_student = user.is_authenticated and user.is_student
    is_teacher = user.is_authenticated and user.is_teacher
    
    if is_student:
        # Students can only export their own attendance
        if user.id != student.id:
            messages.error(request, "You can only export your own attendance.")
            return redirect('SubjectList')
        # Check if student is enrolled in subject
        has_access, redirect_response = check_subject_access(request, subject)
        if not has_access:
            return redirect_response
    elif is_teacher:
        # Teachers must be assigned to subject
        has_access, redirect_response = check_subject_access(request, subject, require_teacher=True)
        if not has_access:
            return redirect_response
    else:
        # Admin/other roles - check general permissions
        has_access, redirect_response = check_subject_access(request, subject)
        if not has_access:
            return redirect_response
    # ===== END AUTHORIZATION CHECK =====
    
    # Role check
    role_name = request.user.role_name
    is_teacher = role_name == 'teacher'
    is_staff = request.user.is_staff
    is_student_owner = request.user.id == student_id

    if not (is_teacher or is_staff or is_student_owner):
        messages.error(request, "You don't have permission to export this attendance data.")
        return redirect('student_subject_attendance', subject_id=subject_id, student_id=student_id)

    # Get semester
    selected_semester_id = request.GET.get('semester')
    if selected_semester_id and selected_semester_id != 'None':
        selected_semester = get_object_or_404(Semester, id=selected_semester_id)
    else:
        selected_semester = Semester.objects.filter(
            start_date__lte=timezone.now().date(),
            end_date__gte=timezone.now().date()
        ).first()

    if not selected_semester:
        messages.error(request, "No active semester found.")
        return redirect('student_subject_attendance', subject_id=subject_id, student_id=student_id)
    
    # Date range
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        # Default to semester date range if no specific dates provided
        start_date = selected_semester.start_date
        end_date = min(selected_semester.end_date, timezone.localdate())

    # Get attendance records
    attendance_records = Attendance.objects.filter(
        student=student,
        subject=subject,
        date__range=(start_date, end_date)
    ).select_related('status', 'teacher').order_by('date')

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Attendance"
    
    # Add title and information rows
    ws.append([f"Student Attendance Report: {student.profile.first_name} {student.profile.last_name}"])
    ws.append([f"Subject: {subject.subject_name}"])
    ws.append([f"Semester: {selected_semester.semester_name} - {selected_semester.start_date.year if selected_semester.start_date.year == selected_semester.end_date.year else f'{selected_semester.start_date.year}-{selected_semester.end_date.year}'}"])
    ws.append([f"Date Range: {start_date.strftime('%m/%d/%Y')} to {end_date.strftime('%m/%d/%Y')}"])
    ws.append([]) # Empty row for spacing
    
    # Make title rows bold
    for row in range(1, 5):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Header row
    header = ['Date', 'Status', 'Remark', 'Graded', 'Points', 'Marked By']
    ws.append(header)
    for col in range(1, len(header) + 1):
        ws.cell(row=6, column=col).font = Font(bold=True)
    
    # Status counts for summary
    status_counts = {}
    total_points = 0
    
    # Data rows
    for record in attendance_records:
        status = record.status.status if record.status else '--'
        
        # Count statuses for summary
        if status in status_counts:
            status_counts[status] += 1
        else:
            status_counts[status] = 1
            
        # Calculate points if graded - use actual teacher-configured points
        points = 0
        if record.graded and record.status and record.teacher:
            teacher_points = TeacherAttendancePoints.objects.filter(
                teacher=record.teacher,
                status=record.status
            ).first()
            if teacher_points:
                points = float(teacher_points.points)
            total_points += points
        
        teacher_name = f"{record.teacher.profile.first_name} {record.teacher.profile.last_name}" if record.teacher else "N/A"
        
        row = [
            record.date.strftime('%Y-%m-%d'),
            status,
            record.remark or "—",
            "Yes" if record.graded else "No",
            points,
            teacher_name
        ]
        ws.append(row)
    
    # Add summary section
    ws.append([]) # Empty row
    ws.append(["Summary"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    
    for status, count in status_counts.items():
        ws.append([f"{status}: {count}"])
    
    if total_points > 0:
        ws.append([])
        ws.append(["Total Graded Points:", total_points])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
        ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"attendance_{student.profile.first_name}_{student.profile.last_name}_{subject.subject_name}_{selected_semester.semester_name}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save workbook to response
    wb.save(response)
    return response
