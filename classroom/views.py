# pyrefly: ignore [missing-import]
from django.shortcuts import render, get_object_or_404, redirect
# pyrefly: ignore [missing-import]
from rest_framework.viewsets import ModelViewSet
from .serializers import TeacherAttendanceSerializer, ClassroomModeSerializer
from .utils import format_duration, format_time, duration_to_minutes, get_matching_dates, build_attendance_report_data
import random
from subject.models import Subject, Schedule
from course.models import SubjectEnrollment, Semester
from accounts.models import CustomUser
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import action
from .models import *
from rest_framework.response import Response
from django.utils.timezone import localtime, now
from django.db.models import Sum, F, ExpressionWrapper, DurationField, Q
from django.db.models.functions import TruncDate
from rest_framework import status
from datetime import datetime, timedelta, date
from calendars.models import Holiday
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.core.files.base import ContentFile
import base64
import json
from django.contrib.auth.decorators import login_required
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import io
from django.utils import timezone
from django.utils.timezone import make_aware, get_current_timezone
from django.utils.timezone import localtime
from rest_framework.permissions import IsAuthenticated
from django.contrib import messages
from django.db import transaction

# Create your views here.

def format_total_time(duration):
    if duration is None:
        return "N/A"
    total_seconds = duration.total_seconds()
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours}h {minutes}m"


class TeacherAttendanceViewSet(ModelViewSet):
    serializer_class = TeacherAttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.serializer_class.Meta.model.objects.all()

    def list(self, request, *args, **kwargs):
        view_type = request.query_params.get('view_type', 'daily')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()

        today = datetime.now().date()
        if view_type == 'daily':
            if not start_date:
                start_date = today
            if not end_date:
                end_date = start_date
        elif view_type == 'weekly':
            if not start_date:
                start_date = today - timedelta(days=today.weekday())
            if not end_date:
                end_date = start_date + timedelta(days=6)
        elif view_type == 'monthly':
            if not start_date:
                start_date = today.replace(day=1)
            if not end_date:
                end_date = (start_date + timedelta(days=31)).replace(day=1) - timedelta(days=1)

        # Generate the full date range
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]

        # Fetch holidays in the date range
        holidays = Holiday.objects.filter(date__gte=start_date, date__lte=end_date)
        holiday_map = {holiday.date: (holiday.title, holiday.holiday_type) for holiday in holidays}

        # Fetch attendance data
        attendance_data = Teacher_Attendance.objects.filter(
            time_started__date__gte=start_date,
            time_started__date__lte=end_date,
        ).annotate(
            day=TruncDate('time_started'),
            duration=ExpressionWrapper(F('time_ended') - F('time_started'), output_field=DurationField()),
        )

        attendance_map = {
            (att.teacher_id, att.subject_id, att.day): att for att in attendance_data
        }

        # Get current semester based on date range
        today_date = datetime.now().date()
        current_semester = Semester.objects.filter(
            start_date__lte=today_date,
            end_date__gte=today_date
        ).first()

        # Fetch schedules filtered by current semester
        if current_semester:
            schedules = Schedule.objects.filter(semester=current_semester).prefetch_related('subject__assign_teacher', 'subject__substitute_teacher')
        else:
            schedules = Schedule.objects.none()

        grouped_data = {}
        for schedule in schedules:
            subject = schedule.subject
            teacher = subject.active_teacher
            if not teacher:
                continue

            teacher_name = f"{teacher.first_name} {teacher.last_name}"
            subject_name = subject.subject_name

            if teacher_name not in grouped_data:
                grouped_data[teacher_name] = {}

            if subject_name not in grouped_data[teacher_name]:
                grouped_data[teacher_name][subject_name] = []


            schedule_dates = [
                current_date for current_date in date_range
                if current_date.strftime('%a') in schedule.days_of_week
            ]

            for current_date in date_range:
                is_holiday = current_date in holiday_map
                attendance = attendance_map.get((teacher.id, subject.id, current_date))
                today = datetime.now().date()

                attendance_id = None
                screenshots = []
                if attendance:
                    screenshots = Screenshot.objects.filter(teacher_attendance=attendance).values_list('image', flat=True)
                    attendance_id = attendance.id

                if is_holiday:
                    holiday_title, holiday_type = holiday_map[current_date]
                    if schedule.schedule_type in ['Regular', 'Build in']:
                        # Calculate total time for the scheduled period
                        schedule_start = datetime.combine(current_date, schedule.schedule_start_time)
                        schedule_end = datetime.combine(current_date, schedule.schedule_end_time)
                        total_time = format_total_time(schedule_end - schedule_start)
                    else:
                        total_time = "N/A"

                    grouped_data[teacher_name][subject_name].append({
                        "attendance_id": attendance_id,
                        "date": current_date,
                        "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                        "schedule_type": schedule.schedule_type,
                        "status": f"Holiday ({holiday_type}) - {holiday_title}",
                        "time_started": None,
                        "time_ended": None,
                        "total_time": total_time,
                        "screenshots": list(screenshots),
                    })
                elif current_date in schedule_dates:
                    if attendance:
                        grouped_data[teacher_name][subject_name].append({
                            "attendance_id": attendance_id,
                            "date": current_date,
                            "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                            "schedule_type": schedule.schedule_type,
                            "status": "Present",
                            "time_started": attendance.time_started,
                            "time_ended": attendance.time_ended,
                            "total_time": format_total_time(attendance.duration),
                            "screenshots": list(screenshots),
                        })
                    else:
                        if current_date < today:
                            # Past scheduled day with no attendance recorded
                            grouped_data[teacher_name][subject_name].append({
                                "attendance_id": attendance_id,
                                "date": current_date,
                                "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                                "schedule_type": schedule.schedule_type,
                                "status": "Absent",
                                "time_started": None,
                                "time_ended": None,
                                "total_time": "N/A",
                                "screenshots": [],
                            })
                        else:
                            # Future scheduled day with no attendance
                            grouped_data[teacher_name][subject_name].append({
                                "attendance_id": attendance_id,
                                "date": current_date,
                                "schedule": f"{schedule.schedule_start_time} - {schedule.schedule_end_time}",
                                "schedule_type": schedule.schedule_type,
                                "status": "No Class",
                                "time_started": None,
                                "time_ended": None,
                                "total_time": "N/A",
                                "screenshots": [],
                            })
                else:
                    # No schedule for this date
                    grouped_data[teacher_name][subject_name].append({
                        "attendance_id": attendance_id,
                        "date": current_date,
                        "schedule": "No Schedule",
                        "schedule_type": schedule.schedule_type,
                        "status": "No Schedule",
                        "time_started": None,
                        "time_ended": None,
                        "total_time": "N/A",
                        "screenshots": [],
                    })

        return Response(grouped_data)

    def export_to_excel(self, request, *args, **kwargs):
        current_date = datetime.now().date()

        # Determine the current semester
        current_semester = Semester.objects.filter(start_date__lte=current_date, end_date__gte=current_date).first()
        current_semester_name = current_semester.semester_name if current_semester else "N/A"

        view_type = request.query_params.get('view_type', 'daily')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Call the existing list method to get the grouped data
        response = self.list(request, *args, **kwargs)
        grouped_data = response.data

        # Create an Excel workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Teacher Attendance"

        # Write headers
        headers = [
            "Teacher Name", "Subject Name", "Subject Code", "Current Semester",
            "Date", "Schedule", "Schedule Type", "Status", "Time Started", "Time Ended", "Total Time"
        ]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Populate data
        row_num = 2
        for teacher, subjects in grouped_data.items():
            for subject, attendance_list in subjects.items():

                subject_obj = Subject.objects.filter(subject_name=subject).first()
                subject_code = subject_obj.subject_code if subject_obj else "N/A"

                for attendance in attendance_list:
                    # Ensure datetime fields are naive
                    date = attendance["date"]
                    time_started = attendance["time_started"]
                    time_ended = attendance["time_ended"]

                    # Convert datetime values to naive (remove timezone)
                    if isinstance(date, datetime):
                        date = date.replace(tzinfo=None).strftime('%Y-%m-%d')
                    if isinstance(time_started, datetime):
                        time_started = localtime(time_started).replace(tzinfo=None).strftime('%H:%M:%S')
                    if isinstance(time_ended, datetime):
                        time_ended = localtime(time_ended).replace(tzinfo=None).strftime('%H:%M:%S')

                    sheet.cell(row=row_num, column=1, value=teacher)
                    sheet.cell(row=row_num, column=2, value=subject)
                    sheet.cell(row=row_num, column=3, value=subject_code)
                    sheet.cell(row=row_num, column=4, value=current_semester_name)
                    sheet.cell(row=row_num, column=5, value=date)
                    sheet.cell(row=row_num, column=6, value=attendance["schedule"])
                    sheet.cell(row=row_num, column=7, value=attendance["schedule_type"])
                    sheet.cell(row=row_num, column=8, value=attendance["status"])
                    sheet.cell(row=row_num, column=9, value=time_started)
                    sheet.cell(row=row_num, column=10, value=time_ended)
                    sheet.cell(row=row_num, column=11, value=attendance["total_time"])
                    row_num += 1

        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            sheet.column_dimensions[column_letter].width = 20

        # Create HTTP response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename=Teacher_Attendance_{view_type}.xlsx"
        workbook.save(response)
        return response

    @action(detail=True, methods=['post'], url_path='start-class')
    def start_class(self, request, pk=None):
        subject_id = pk
        user = request.user

        # Ensure the subject exists (add your Subject model import)
        try:
            subject = Subject.objects.get(pk=subject_id)
        except Subject.DoesNotExist:
            return Response({'error': 'Subject not found'}, status=404)

        # Check if the user is the assigned teacher or substitute teacher
        if subject.assign_teacher != user and not (subject.substitute_teacher == user and subject.allow_substitute_teacher):
            return Response({'error': 'You are not assigned to this subject.'}, status=403)

        # Validate if the current time is within the schedule or up to 15 minutes early
        current_time = localtime(now())
        current_day = current_time.strftime('%a')

        # Get current semester
        today_date = current_time.date()
        current_semester = Semester.objects.filter(
            start_date__lte=today_date,
            end_date__gte=today_date
        ).first()

        if not current_semester:
            return Response({'error': 'No active semester found.'}, status=400)

        # Get schedules for today filtered by current semester
        schedules = subject.schedules.filter(
            days_of_week__icontains=current_day,
            semester=current_semester
        )
        if not schedules.exists():
            return Response({'error': f'No schedules found for today ({current_day}) in the current semester.'}, status=400)

        # Find the appropriate schedule for the current time
        valid_schedule = None
        for schedule in schedules:
            # Allow starting up to 15 minutes early
            early_start_time = (datetime.combine(current_time.date(), schedule.schedule_start_time) - timedelta(minutes=15)).time()

            # Check if current time is within the schedule or up to 15 minutes early
            if (early_start_time <= current_time.time() <= schedule.schedule_end_time):
                valid_schedule = schedule
                break

        if not valid_schedule:
            return Response({'error': f'You can only start the class during your scheduled time or up to 15 minutes early.'}, status=400)

        # Always record the actual time the teacher started the class.
        # Previously, early starts used the scheduled start time (future),
        # which caused the frontend timer to display negative elapsed time.
        record_time = current_time

        # Atomically check for an existing active session and create one if absent
        with transaction.atomic():
            existing_attendance = Teacher_Attendance.objects.select_for_update().filter(
                subject=subject,
                teacher=user,
                is_active=True
            ).first()

            if existing_attendance:
                serializer = self.get_serializer(existing_attendance)
                return Response({
                    'message': 'Class is already active.',
                    'attendance_id': existing_attendance.id,
                    'scheduled_end_time': valid_schedule.schedule_end_time.strftime('%H:%M:%S'),
                    'time_started': existing_attendance.time_started.isoformat(),
                    'data': serializer.data
                }, status=200)

            teacher_attendance = Teacher_Attendance.objects.create(
                subject=subject,
                teacher=user,
                time_started=record_time,
                is_active=True
            )

        # Schedule Celery task to auto-end class at scheduled end time
        from classroom.tasks import auto_end_class

        # Calculate time until scheduled end
        scheduled_end_datetime = datetime.combine(current_time.date(), valid_schedule.schedule_end_time)
        scheduled_end_datetime = make_aware(scheduled_end_datetime, get_current_timezone())

        # Schedule the task to run at the scheduled end time
        task = auto_end_class.apply_async(
            args=[teacher_attendance.id],
            eta=scheduled_end_datetime
        )

        # Store the task ID so we can cancel it if needed
        teacher_attendance.celery_task_id = task.id
        teacher_attendance.save()

        serializer = self.get_serializer(teacher_attendance)
        return Response({
            'message': 'Class started successfully!',
            'attendance_id': teacher_attendance.id,
            'scheduled_end_time': valid_schedule.schedule_end_time.strftime('%H:%M:%S'),
            'time_started': teacher_attendance.time_started.isoformat(),
            'data': serializer.data
        }, status=201)


    @action(detail=True, methods=['post'], url_path='end-class')
    def end_class(self, request, pk=None):
        try:
            # Get all active attendance records (in case of duplicates from multiple clicks)
            active_attendances = Teacher_Attendance.objects.filter(
                subject__id=pk,
                teacher=request.user,
                is_active=True
            )

            if not active_attendances.exists():
                return Response({'error': 'Active classroom mode not found.'}, status=404)

            # Use the first one for schedule/time processing
            teacher_attendance = active_attendances.first()

            current_time = localtime(now())
            current_day = current_time.strftime('%a')

            # Get current semester
            today_date = current_time.date()
            current_semester = Semester.objects.filter(
                start_date__lte=today_date,
                end_date__gte=today_date
            ).first()

            # Get the schedule for today to find the scheduled end time
            subject = teacher_attendance.subject
            schedules = subject.schedules.filter(
                days_of_week__icontains=current_day
            )

            # Add semester filter if current semester exists
            if current_semester:
                schedules = schedules.filter(semester=current_semester)

            # Match the schedule allowing for 15-min early starts
            scheduled_end_time = None
            if schedules.exists():
                start_time = localtime(teacher_attendance.time_started).time()
                for schedule in schedules:
                    early_start = (datetime.combine(current_time.date(), schedule.schedule_start_time) - timedelta(minutes=15)).time()
                    if early_start <= start_time <= schedule.schedule_end_time:
                        scheduled_end_time = schedule.schedule_end_time
                        break

            # Determine the end time to save
            if scheduled_end_time:
                # If current time is after scheduled end time, use scheduled end time
                # Otherwise, use actual current time
                if current_time.time() > scheduled_end_time:
                    # Cap at scheduled end time
                    end_time = datetime.combine(current_time.date(), scheduled_end_time)
                    end_time = localtime(make_aware(end_time))
                else:
                    # Use actual time (ending early)
                    end_time = current_time
            else:
                # No schedule found, use actual time
                end_time = current_time

            # End all active attendance records (cleanup duplicates if any)
            from celery.result import AsyncResult
            from lms.celery import app as celery_app
            for attendance in active_attendances:
                # Revoke scheduled Celery task for each record
                if attendance.celery_task_id:
                    try:
                        AsyncResult(attendance.celery_task_id, app=celery_app).revoke(terminate=True)
                    except Exception as e:
                        print(f"Could not cancel task {attendance.celery_task_id}: {str(e)}")
                attendance.time_ended = end_time
                attendance.is_active = False
                attendance.manual_ended = True
                attendance.celery_task_id = None
                attendance.save()

            return Response({'message': 'Classroom mode ended successfully.'}, status=200)
        except Exception as e:
            return Response({'error': f'An error occurred: {str(e)}'}, status=500)



    @action(detail=True, methods=['get'], url_path='current-state')
    def current_state(self, request, pk=None):
        try:
            teacher_attendance = Teacher_Attendance.objects.filter(
                subject__id=pk, teacher=request.user, is_active=True
            ).order_by('-time_started').first()
            if teacher_attendance:
                return Response({
                    'is_active': True,
                    'attendance_id': teacher_attendance.id,
                    'time_started': teacher_attendance.time_started.isoformat()
                }, status=200)
            return Response({'is_active': False}, status=200)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


    @action(detail=True, methods=['get'], url_path='get-end-time')
    def get_end_time(self, request, pk=None):
        """Fetches the scheduled end time for a class based on subject ID."""
        try:
            subject = Subject.objects.get(pk=pk)  # Fetch by subject ID
            current_time = localtime(now())
            current_day = current_time.strftime('%a')

            # Get current semester
            today_date = current_time.date()
            current_semester = Semester.objects.filter(
                start_date__lte=today_date,
                end_date__gte=today_date
            ).first()

            if not current_semester:
                return Response({'error': 'No active semester found.'}, status=400)

            # Get schedules for today filtered by current semester
            schedules = subject.schedules.filter(
                days_of_week__icontains=current_day,
                semester=current_semester
            )
            if schedules.exists():
                latest_end_time = max(schedule.schedule_end_time for schedule in schedules)
                return Response({"end_time": latest_end_time.strftime('%H:%M:%S')}, status=200)

            return Response({"error": "No schedule found for today."}, status=404)

        except Subject.DoesNotExist:
            return Response({"error": "Subject not found."}, status=404)


class ClassroomModeViewSet(ModelViewSet):
    serializer_class = ClassroomModeSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return self.serializer_class.Meta.model.objects.all()
    
    @action(detail=False, methods=['post'], url_path='toggle-mode')
    def toggle_classroom_mode(self, request):
        classroom_mode_instance, created = Classroom_mode.objects.get_or_create(id=1)  
        classroom_mode_instance.is_classroom_mode = not classroom_mode_instance.is_classroom_mode
        classroom_mode_instance.save()

        return Response(
            {"is_classroom_mode": classroom_mode_instance.is_classroom_mode},
            status=status.HTTP_200_OK
        )

@login_required
def enter_classroom_mode_view(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    classroom_mode_instance, _ = Classroom_mode.objects.get_or_create(subject=subject)
    classroom_mode_instance.is_classroom_mode = True
    classroom_mode_instance.save()
    return redirect('classroom_mode', pk=subject.id)

@login_required
def exit_classroom_mode_view(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    try:
        classroom_mode_instance = Classroom_mode.objects.get(subject=subject)
        classroom_mode_instance.is_classroom_mode = False
        classroom_mode_instance.save()
    except Classroom_mode.DoesNotExist:
        pass  # No need to do anything if the mode was never started
    return redirect('material-list', id=subject.id)

@login_required
def lucky_draw(request, subject_id):
    """
    View for conducting a lucky draw for students enrolled in a specific subject
    within the current semester. Allows manual inclusion of a student and
    excluding previously winning students.
    """
    # Get the current semester based on today's date
    current_date = timezone.localtime(timezone.now()).date()
    current_semester = Semester.objects.filter(start_date__lte=current_date, end_date__gte=current_date).first()

    if not current_semester:
        return JsonResponse({
            'error': 'No active semester is found.',
            'students': [],
            'winner': None,
        }, status=404)

    # Get the subject
    subject = get_object_or_404(Subject, id=subject_id)

    # Get all students enrolled in the subject for the current semester
    students = list(SubjectEnrollment.objects.filter(
        subject=subject,
        semester=current_semester,
        status='enrolled'
    ).values('student__id', 'student__first_name', 'student__last_name'))

    # Optionally exclude previously winning students
    exclude_winners = request.GET.get('exclude_winners', 'false').lower() == 'true'
    if exclude_winners:
        # Fetch previously winning students (this can be fetched from a model or session)
        # For simplicity, assume `winning_students` is stored in the session
        winning_students = request.session.get('winning_students', [])
        students = [student for student in students if str(student['student__id']) not in winning_students]

    # Allow manual inclusion of a student
    manual_student_id = request.GET.get('manual_student_id', None)
    if manual_student_id:
        student = CustomUser.objects.filter(id=manual_student_id).values(
            'id', 'first_name', 'last_name'
        ).first()
        if student and student not in students:
            students.append({
                'student__id': student['id'],
                'student__first_name': student['first_name'],
                'student__last_name': student['last_name']
            })

    # Randomly select a winner
    winner = random.choice(students) if students else None

    # Save the winner to session if one exists
    if winner:
        winning_students = request.session.get('winning_students', [])
        winning_students.append(str(winner['student__id']))
        request.session['winning_students'] = winning_students

    return JsonResponse({
        'subject': subject.subject_name,
        'students': students,
        'winner': winner,
    })

@login_required
def lucky_draw_page(request, subject_id):
    subject = get_object_or_404(Subject, id=subject_id)
    return render(request, 'classroom/lucky_draw.html', {'subject': subject})

@login_required
def classroom_dashboard(request):
    return render(request, 'classroom/Classroom_dashboard.html')

@login_required
def reset_lucky_draw(request, subject_id):
    """
    View to reset the lucky draw by clearing the list of previously winning students.
    """
    subject = get_object_or_404(Subject, id=subject_id)

    # Clear the session data for winning students
    if 'winning_students' in request.session:
        del request.session['winning_students']

    return JsonResponse({
        'message': f"The lucky draw for {subject.subject_name} has been reset.",
        'status': 'success'
    })

@login_required
def teacher_attendance(request):
    subject = Subject.objects.all()
    return render(request, 'timesheet/teacher_attendance_list.html',{'subject': subject})


@login_required
def teacher_attendance_details(request, id):
    attendance_details = Teacher_Attendance.objects.filter(subject=id)
    
    subject_id = attendance_details.first().subject.id

    # Calculate total time for each attendance record
    for attendance in attendance_details:
        if attendance.time_ended:
            total_seconds = (attendance.time_ended - attendance.time_started).total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            attendance.total_time = f"{hours}h {minutes}m"
        else:
            attendance.total_time = "N/A"

    return render(request, 'timesheet/teacher_attendance_details.html', {
        'attendance_details': attendance_details,
        "subject_id": subject_id,
    })


@login_required
def teacher_attendance_details_per_day(request, id):
    date_selected = request.GET.get("date")  # Get date from the URL

    # Validate date format
    try:
        selected_date = datetime.strptime(date_selected, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return HttpResponse("Invalid date format", status=400)

    # Filter attendance records based on subject and date
    attendance_details = Teacher_Attendance.objects.filter(
        subject_id=id,
        time_started__date=selected_date
    )

    # Calculate total time for each attendance record
    for attendance in attendance_details:
        if attendance.time_ended:
            total_seconds = (attendance.time_ended - attendance.time_started).total_seconds()
            hours, remainder = divmod(total_seconds, 3600)
            minutes, _ = divmod(remainder, 60)
            attendance.total_time = f"{hours}h {minutes}m"
        else:
            attendance.total_time = "N/A"

    return render(request, 'timesheet/teacher_attendance_details.html', {
        'attendance_details': attendance_details,
        'selected_date': selected_date
    })



@login_required
def view_screenshots(request, id):
    attendance = get_object_or_404(Teacher_Attendance, id=id)
    screenshots = Screenshot.objects.filter(teacher_attendance=attendance)
    subject_id = attendance.subject.id if attendance.subject else None

    return render(request, "timesheet/view_all_screenshot.html", {
        "attendance": attendance,
        "screenshots": screenshots,
        "subject_id": subject_id,
    })


@login_required
def view_screenshots_per_date(request, subject_id, selected_date):
    """
    View screenshots only for the given subject and date.
    """
    try:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()  

    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Filter attendance records for the selected subject and date
    attendance_records = Teacher_Attendance.objects.filter(
        subject_id=subject_id,
        time_started__date=selected_date
    )

    # Get all screenshots related to the filtered attendance records
    screenshots = Screenshot.objects.filter(teacher_attendance__in=attendance_records)

    return render(request, "timesheet/view_screenshot_per_date.html", {
        "attendance_records": attendance_records,
        "screenshots": screenshots,
        "selected_date": selected_date,
        "subject_id": subject_id,
    })


@login_required
def export_screenshots_pdf(request, subject_id, selected_date):
    """
    Generate a PDF with all screenshots and timestamps for the selected date.
    """
    try:
        selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get subject and teacher information
    subject = get_object_or_404(Subject, id=subject_id)
    attendance_records = Teacher_Attendance.objects.filter(subject_id=subject_id, time_started__date=selected_date)
    
    # Get the teacher name (assuming the first attendance record is valid)
    teacher_name = attendance_records.first().teacher.get_full_name() if attendance_records.exists() else "Unknown"

    # Get all screenshots
    screenshots = Screenshot.objects.filter(teacher_attendance__in=attendance_records)

    # Create a PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="screenshots_{selected_date}.pdf"'

    # Create a PDF object
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter  

    # Title & Header
    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(50, height - 50, f"Screenshots for {selected_date}")

    pdf.setFont("Helvetica", 12)
    pdf.drawString(50, height - 70, f"Subject: {subject.subject_name}")
    pdf.drawString(50, height - 90, f"Teacher: {teacher_name}")

    y_position = height - 140  

    for screenshot in screenshots:
        if y_position < 250:  
            pdf.showPage()
            pdf.setFont("Helvetica-Bold", 16)
            pdf.drawString(50, height - 50, f"Screenshots for {selected_date}")
            pdf.setFont("Helvetica", 12)
            pdf.drawString(50, height - 70, f"Subject: {subject.subject_name}")
            pdf.drawString(50, height - 90, f"Teacher: {teacher_name}")
            y_position = height - 140  

        # Draw timestamp above the image
        pdf.setFont("Helvetica", 12)
        pdf.drawString(50, y_position, f"Taken on: {screenshot.timestamp.strftime('%Y-%m-%d %H:%M:%S')}")

        # Load and fit image properly
        try:
            image = ImageReader(screenshot.image.path)
            img_width = width - 100  
            img_height = 200  
            pdf.drawImage(image, 50, y_position - img_height - 10, width=img_width, height=img_height, preserveAspectRatio=True)
            y_position -= img_height + 40  
        except Exception as e:
            pdf.drawString(50, y_position - 20, "Error loading image")
            y_position -= 40

    # Save PDF
    pdf.save()
    buffer.seek(0)
    response.write(buffer.read())
    return response

@login_required
def save_screenshot(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({"error": "Invalid JSON data"}, status=400)

        image_data = data.get("image")
        attendance_id = data.get("attendance_id")

        if not image_data or not attendance_id:
            return JsonResponse({"error": "Missing image data or attendance_id"}, status=400)

        # Ensure attendance exists
        try:
            attendance = Teacher_Attendance.objects.get(id=attendance_id)
        except Teacher_Attendance.DoesNotExist:
            return JsonResponse({"error": "Attendance record not found"}, status=404)

        # Fix base64 decoding: Remove data URI prefix
        if "," in image_data:
            format_info, imgstr = image_data.split(";base64,")
        else:
            return JsonResponse({"error": "Invalid base64 format"}, status=400)

        ext = format_info.split("/")[-1]  

        # Decode and save image
        try:
            image_file = ContentFile(base64.b64decode(imgstr), name=f"screenshot_{now().strftime('%Y%m%d_%H%M%S')}.{ext}")
            screenshot = Screenshot.objects.create(teacher_attendance=attendance, image=image_file)
        except Exception as e:
            return JsonResponse({"error": f"Failed to decode image: {str(e)}"}, status=400)

        return JsonResponse({"message": "Screenshot saved successfully", "image_url": screenshot.image.url})

    return JsonResponse({"error": "Invalid request"}, status=400)



@login_required
def teacher_timesheet_report(request):
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    teacher_id = request.GET.get("teacher_id")
    semester_id = request.GET.get("semester_id")

    today = date.today()
    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get all semesters for the filter dropdown
    semesters = Semester.objects.all().order_by('-start_date')
    
    # Get current semester based on date range
    today_date = date.today()
    current_semester = Semester.objects.filter(
        start_date__lte=today_date,
        end_date__gte=today_date
    ).first()
    
    # Determine which semester to use for filtering
    selected_semester = None
    if semester_id:
        try:
            selected_semester = Semester.objects.get(id=int(semester_id))
        except (ValueError, TypeError, Semester.DoesNotExist):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    # Get all teachers for the filter dropdown
    teachers = CustomUser.objects.filter(
        primary_teacher__isnull=False
    ).distinct().order_by('first_name', 'last_name')

    # Get all subjects and assigned teachers
    subjects_query = Subject.objects.values(
        "id", "subject_name", "assign_teacher__id",
        "assign_teacher__first_name", "assign_teacher__last_name"
    )
    
    # Apply teacher filter if provided
    if teacher_id:
        try:
            # Convert teacher_id to integer for database filtering
            teacher_id_int = int(teacher_id)
            subjects_query = subjects_query.filter(assign_teacher__id=teacher_id_int)
        except (ValueError, TypeError):
            # If conversion fails, don't filter
            pass

    # Process schedule data with semester filter
    if teacher_id:
        try:
            teacher_id_int = int(teacher_id)
            schedule_data = Schedule.objects.filter(subject__assign_teacher__id=teacher_id_int)
        except (ValueError, TypeError):
            schedule_data = Schedule.objects.all()
    else:
        schedule_data = Schedule.objects.all()

    # Apply semester filter to schedules
    if selected_semester:
        schedule_data = schedule_data.filter(semester=selected_semester)

    # Orbit Program subjects (COIL / HALI / CTE) are tracked outside the
    # regular timesheet rhythm — they have their own program-specific
    # schedules and attendance flows — so leave them out of the teacher
    # timesheet report.
    schedule_data = schedule_data.exclude(
        Q(subject__is_coil=True) | Q(subject__is_hali=True)
    )

    final_report_data, all_dates_list = build_attendance_report_data(schedule_data, start_date, end_date, teacher_id)

    timesheet_stats = _compute_timesheet_stats(final_report_data)

    return render(request, 'timesheet/teacher_timesheet_report.html', {
        "report_data": final_report_data,
        "all_dates": all_dates_list,
        "start_date": start_date.strftime("%Y-%m-%d") if start_date else "",
        "end_date": end_date.strftime("%Y-%m-%d") if end_date else "",
        "teachers": teachers,
        "selected_teacher_id": teacher_id,
        "semesters": semesters,
        "selected_semester_id": semester_id if semester_id else (selected_semester.id if selected_semester else None),
        "current_semester": current_semester,
        "timesheet_stats": timesheet_stats,
    })


def _compute_timesheet_stats(report_data):
    """Aggregate per-subject report rows into the summary tiles the
    timesheet report template renders at the top of the page.
    """
    scheduled_minutes = 0
    actual_minutes = 0
    total_entries = 0
    sessions_with_attendance = 0
    absent = 0
    on_time = 0
    late = 0

    for row in (report_data or []):
        scheduled_minutes += duration_to_minutes(row.get("total_scheduled_time"))
        actual_minutes += duration_to_minutes(row.get("total_attendance_time"))

        for entry in row.get("date_entries", []):
            total_entries += 1
            if entry.get("has_attendance"):
                sessions_with_attendance += 1
                budget = entry.get("budget_start")
                actual = entry.get("actual_start")
                if budget and actual and budget != "-" and actual != "-":
                    if actual > budget:
                        late += 1
                    else:
                        on_time += 1
                else:
                    on_time += 1
            else:
                absent += 1

    def fmt(mins):
        if mins < 0:
            mins = -mins
        return f"{mins // 60}h {mins % 60}m"

    variance_minutes = scheduled_minutes - actual_minutes
    attendance_rate = round((sessions_with_attendance / total_entries) * 100) if total_entries else 0
    on_time_pct = round((on_time / sessions_with_attendance) * 100) if sessions_with_attendance else 0

    return {
        "scheduled_hours": fmt(scheduled_minutes),
        "actual_hours": fmt(actual_minutes),
        "variance_hours": fmt(variance_minutes),
        "variance_negative": variance_minutes < 0,
        "total_entries": total_entries,
        "attendance_rate": attendance_rate,
        "on_time": on_time,
        "late": late,
        "on_time_pct": on_time_pct,
        "absent": absent,
    }
    
  
   
@login_required
def export_teacher_attendance_excel(request):
    """
    Export teacher attendance report data to Excel based on date range and teacher selection
    """
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    teacher_id = request.GET.get("teacher_id")
    semester_id = request.GET.get("semester_id")

    today = date.today()
    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get current semester based on date range
    today_date = date.today()
    current_semester = Semester.objects.filter(
        start_date__lte=today_date,
        end_date__gte=today_date
    ).first()
    
    # Determine which semester to use for filtering
    selected_semester = None
    if semester_id:
        try:
            selected_semester = Semester.objects.get(id=int(semester_id))
        except (ValueError, TypeError, Semester.DoesNotExist):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    # Get all teachers for the filter dropdown
    teachers = CustomUser.objects.filter(
        primary_teacher__isnull=False
    ).distinct().order_by('first_name', 'last_name')

    # Get all subjects and assigned teachers
    subjects_query = Subject.objects.values(
        "id", "subject_name", "assign_teacher__id",
        "assign_teacher__first_name", "assign_teacher__last_name"
    )
    
    # Apply teacher filter if provided
    if teacher_id:
        try:
            # Convert teacher_id to integer for database filtering
            teacher_id_int = int(teacher_id)
            subjects_query = subjects_query.filter(assign_teacher__id=teacher_id_int)
        except (ValueError, TypeError):
            # If conversion fails, don't filter
            pass

    # Process schedule data
    if teacher_id:
        try:
            teacher_id_int = int(teacher_id)
            schedule_data = Schedule.objects.filter(subject__assign_teacher__id=teacher_id_int)
        except (ValueError, TypeError):
            schedule_data = Schedule.objects.all()
    else:
        schedule_data = Schedule.objects.all()
    
    # Apply semester filter to schedules
    if selected_semester:
        schedule_data = schedule_data.filter(semester=selected_semester)

    # Keep the Excel export aligned with the on-screen report — Orbit
    # Program subjects (COIL / HALI / CTE) are out of scope for the
    # regular teacher timesheet.
    schedule_data = schedule_data.exclude(
        Q(subject__is_coil=True) | Q(subject__is_hali=True)
    )

    final_report_data, all_dates_list = build_attendance_report_data(schedule_data, start_date, end_date, teacher_id)

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Teacher Attendance Report"
    
    # Add headers
    headers = [
        "Teacher Name", "Subject", "Date",
        "Scheduled Start", "Scheduled End", "Scheduled Duration",
        "Actual Start", "Actual End", "Actual Duration",
        "Undertime", "Overtime"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    row_num = 2
    for record in final_report_data:
        teacher_name = record["teacher_name"]
        subject_name = record["subject_name"]
        
        # Add detail rows for each date directly (no summary row)
        for entry in record["date_entries"]:
            sheet.cell(row=row_num, column=1, value=teacher_name)
            sheet.cell(row=row_num, column=2, value=subject_name)
            sheet.cell(row=row_num, column=3, value=entry["date"].strftime("%Y-%m-%d"))
            sheet.cell(row=row_num, column=4, value=entry["budget_start"])
            sheet.cell(row=row_num, column=5, value=entry["budget_end"])
            sheet.cell(row=row_num, column=6, value=entry["budget_duration"])
            sheet.cell(row=row_num, column=7, value=entry["actual_start"])
            sheet.cell(row=row_num, column=8, value=entry["actual_end"])
            sheet.cell(row=row_num, column=9, value=entry["actual_duration"])
            
            # Add daily variance to appropriate column
            if entry["daily_variance"]:
                if entry["daily_variance_negative"]:
                    sheet.cell(row=row_num, column=11, value=entry["daily_variance"])  # Overtime
                else:
                    sheet.cell(row=row_num, column=10, value=entry["daily_variance"])  # Undertime
            
            # Highlight rows with attendance
            if entry["has_attendance"]:
                for col in range(1, 12):
                    cell = sheet.cell(row=row_num, column=col)
                    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
            
            row_num += 1
        
        # No empty row for spacing between teachers anymore
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 18
    
    # Create HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Teacher_Attendance_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}.xlsx"
    workbook.save(response)
    return response


@login_required
def teacher_attendance_calendar(request):
    subject_id = request.GET.get("subject_id")  
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else None
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else None
    except ValueError as e:
        return JsonResponse({"error": f"Invalid date format: {e}"}, status=400)

    # Base query
    attendance_data = Teacher_Attendance.objects.filter(
        time_ended__isnull=False,
        subject_id=subject_id
    )

    # Apply date filters if provided
    if start_date:
        attendance_data = attendance_data.filter(time_started__date__gte=start_date)
    if end_date:
        attendance_data = attendance_data.filter(time_started__date__lte=end_date)
    
    attendance_data = (
        attendance_data
        .annotate(date=TruncDate('time_started'))  
        .values('date', 'subject__id')
        .annotate(
            total_attendance_time=Sum(ExpressionWrapper(
                F('time_ended') - F('time_started'), output_field=DurationField()
            ))
        )
    )

    # Format data for FullCalendar
    events = []
    for record in attendance_data:
        total_seconds = int(record["total_attendance_time"].total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, _ = divmod(remainder, 60)
        formatted_time = f"{hours}h {minutes}m"

        events.append({
            "title": formatted_time,  
            "start": record["date"].strftime('%Y-%m-%d'),
        })
    
    return JsonResponse(events, safe=False)



@login_required
def teacher_attendance_calendar_page(request, subject_id):
    # Get date parameters from request
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    # Get subject details
    subject = Schedule.objects.filter(subject__id=subject_id).values(
        'subject__id',
        'subject__subject_name', 
        'subject__assign_teacher__first_name', 
        'subject__assign_teacher__last_name'
    ).first()

    return render(request, 'timesheet/teacher_attendance_calendar.html', {
        "subject": subject,
        "subject_id": subject_id,
        "start_date": start_date,
        "end_date": end_date
    })

@login_required
def find_teacher_attendance(request):
    """
    Find the most recent attendance record for a teacher and subject within a date range.
    Used for linking teacher names in the attendance report to their screenshots.
    """
    teacher_id = request.GET.get('teacher_id')
    subject_id = request.GET.get('subject_id')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    # Parse dates
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        else:
            # Default to 30 days ago if no start date provided
            start_date = datetime.now().date() - timedelta(days=30)
            
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        else:
            # Default to today if no end date provided
            end_date = datetime.now().date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date format'}, status=400)
    
    # Find the most recent attendance record for this teacher and subject
    attendance_query = Teacher_Attendance.objects.filter(
        teacher_id=teacher_id,
        subject_id=subject_id,
        time_started__date__gte=start_date,
        time_started__date__lte=end_date,
        time_ended__isnull=False
    ).order_by('-time_started')
    
    if attendance_query.exists():
        attendance = attendance_query.first()
        return JsonResponse({'attendance_id': attendance.id})
    else:
        return JsonResponse({'attendance_id': None})
    

@login_required
def export_teacher_attendance_summary_excel(request):
    """
    Export teacher attendance summary report to Excel with totals by schedule type (Overload, Build-in, Regular)
    """
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    teacher_id = request.GET.get("teacher_id")
    semester_id = request.GET.get("semester_id")

    today = date.today()
    # Convert start_date and end_date to datetime format
    try:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else today
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date() if end_date else today
    except ValueError:
        return HttpResponse("Invalid date format", status=400)

    # Get current semester based on date range
    today_date = date.today()
    current_semester = Semester.objects.filter(
        start_date__lte=today_date,
        end_date__gte=today_date
    ).first()
    
    # Determine which semester to use for filtering
    selected_semester = None
    if semester_id:
        try:
            selected_semester = Semester.objects.get(id=int(semester_id))
        except (ValueError, TypeError, Semester.DoesNotExist):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    # Get all teachers for the filter dropdown
    teachers = CustomUser.objects.filter(
        primary_teacher__isnull=False
    ).distinct().order_by('first_name', 'last_name')

    # Apply teacher filter if provided
    if teacher_id:
        try:
            teacher_id_int = int(teacher_id)
            teachers = teachers.filter(id=teacher_id_int)
        except (ValueError, TypeError):
            pass

    # Generate a list of all dates in the range
    def daterange(start_date, end_date):
        for n in range(int((end_date - start_date).days) + 1):
            yield start_date + timedelta(n)
    
    # Get all dates in the range
    all_dates = list(daterange(start_date, end_date))

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Teacher Hours Summary"
    
    # Define column widths for fixed columns
    column_widths = {
        1: 30,  # Name
        2: 15,  # Standard Time
        3: 15,  # Built-In Load
        4: 15,  # Regular Load
        5: 15,  # Overload
        6: 15,  # Computed Time
        7: 15,  # Built-In Load
        8: 15,  # Regular Load
        9: 15,  # Overload
        10: 15, # Variance
    }
    
    # Add column widths for date columns
    date_column_start = 11
    for i, _ in enumerate(all_dates):
        column_widths[date_column_start + i] = 10
    
    for col_num, width in column_widths.items():
        sheet.column_dimensions[get_column_letter(col_num)].width = width
    
    # Add headers with merged cells for grouped columns
    sheet.merge_cells('A1:A2')
    sheet.cell(row=1, column=1, value="Name").alignment = Alignment(horizontal="center", vertical="center")
    
    sheet.merge_cells('B1:E1')
    sheet.cell(row=1, column=2, value="Standard Time").alignment = Alignment(horizontal="center")
    
    sheet.merge_cells('F1:I1')
    sheet.cell(row=1, column=6, value="Computed Time").alignment = Alignment(horizontal="center")
    
    sheet.merge_cells('J1:J2')
    sheet.cell(row=1, column=10, value="Variance").alignment = Alignment(horizontal="center", vertical="center")
    
    # Add date range header
    if all_dates:
        date_range_start_col = date_column_start
        date_range_end_col = date_column_start + len(all_dates) - 1
        sheet.merge_cells(f'{get_column_letter(date_range_start_col)}1:{get_column_letter(date_range_end_col)}1')
        sheet.cell(row=1, column=date_range_start_col, value="Daily Attendance").alignment = Alignment(horizontal="center")
        
        # Add individual date headers
        for i, date_obj in enumerate(all_dates):
            col = date_column_start + i
            sheet.cell(row=2, column=col, value=date_obj.strftime("%m/%d")).alignment = Alignment(horizontal="center")
    
    # Second header row for fixed columns
    headers_row2 = [
        "", # Name (already merged)
        "Standard Time", "Built-In Load", "Regular Load", "Overload",
        "Computed Time", "Built-In Load", "Regular Load", "Overload",
        "" # Variance (already merged)
    ]
    
    for col_num, header in enumerate(headers_row2, 1):
        if header:  # Skip empty cells (merged cells)
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.alignment = Alignment(horizontal="center")
    
    # Style all header cells
    for row in range(1, 3):
        for col in range(1, date_column_start + len(all_dates)):
            cell = sheet.cell(row=row, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Process data for each teacher
    row_num = 3
    
    for teacher in teachers:
        teacher_name = f"{teacher.first_name} {teacher.last_name}"
        
        # Initialize counters for different schedule types
        scheduled_times = {
            'standard': timedelta(0),
            'built_in': timedelta(0),
            'regular': timedelta(0),
            'overload': timedelta(0),
        }
        
        actual_times = {
            'standard': timedelta(0),
            'built_in': timedelta(0),
            'regular': timedelta(0),
            'overload': timedelta(0),
        }
        
        # Dictionary to store daily attendance totals
        daily_attendance = {date_obj: timedelta(0) for date_obj in all_dates}
        
        # Get all subjects for this teacher
        subjects = Subject.objects.filter(assign_teacher=teacher)
        
        # Process schedules for each subject
        for subject in subjects:
            schedules = Schedule.objects.filter(subject=subject)
            
            # Apply semester filter to schedules
            if selected_semester:
                schedules = schedules.filter(semester=selected_semester)
            
            for schedule in schedules:
                schedule_type = schedule.schedule_type.lower() if schedule.schedule_type else 'regular'
                if schedule_type not in ['overload', 'build in', 'regular']:
                    schedule_type = 'regular'  # Default to regular if type is not recognized
                
                # Map 'build in' to 'built_in' for consistency
                if schedule_type == 'build in':
                    schedule_type = 'built_in'
                
                # Get matching dates for this schedule
                schedule_days = schedule.days_of_week
                matching_dates = get_matching_dates(schedule_days, start_date, end_date)
                
                # Calculate scheduled time for each matching date
                for match_date in matching_dates:
                    schedule_start = datetime.combine(match_date, schedule.schedule_start_time)
                    schedule_end = datetime.combine(match_date, schedule.schedule_end_time)
                    scheduled_duration = schedule_end - schedule_start
                    
                    # Add to scheduled time for this type
                    scheduled_times[schedule_type] += scheduled_duration
                    scheduled_times['standard'] += scheduled_duration
        
        # Get attendance records for this teacher
        attendance_records = Teacher_Attendance.objects.filter(
            teacher=teacher,
            time_started__date__gte=start_date,
            time_started__date__lte=end_date,
            time_ended__isnull=False
        )
        
        # Group attendance records by subject and date to merge overlapping sessions
        attendance_by_subject_date = {}
        
        for record in attendance_records:
            subject_id = record.subject.id
            record_date = localtime(record.time_started).date()
            key = (subject_id, record_date)
            
            if key not in attendance_by_subject_date:
                attendance_by_subject_date[key] = {
                    'subject': record.subject,
                    'date': record_date,
                    'earliest_start': localtime(record.time_started),
                    'latest_end': localtime(record.time_ended),
                    'schedule_type': None
                }
            else:
                # Update to earliest start and latest end (merge sessions)
                if localtime(record.time_started) < attendance_by_subject_date[key]['earliest_start']:
                    attendance_by_subject_date[key]['earliest_start'] = localtime(record.time_started)
                if localtime(record.time_ended) > attendance_by_subject_date[key]['latest_end']:
                    attendance_by_subject_date[key]['latest_end'] = localtime(record.time_ended)
        
        # Process merged attendance records
        for key, merged_data in attendance_by_subject_date.items():
            subject = merged_data['subject']
            record_date = merged_data['date']
            day_of_week = record_date.strftime("%a")  # Get abbreviated day name (Mon, Tue, etc.)
            
            # Find matching schedule
            matching_schedule = None
            for schedule in Schedule.objects.filter(subject=subject):
                if day_of_week in schedule.days_of_week:
                    matching_schedule = schedule
                    break
            
            # Determine schedule type
            schedule_type = 'regular'  # Default
            if matching_schedule and matching_schedule.schedule_type:
                schedule_type = matching_schedule.schedule_type.lower()
                if schedule_type not in ['overload', 'build in', 'regular']:
                    schedule_type = 'regular'
                
                # Map 'build in' to 'built_in' for consistency
                if schedule_type == 'build in':
                    schedule_type = 'built_in'
            
            # Calculate merged attendance duration (earliest start to latest end)
            attendance_duration = merged_data['latest_end'] - merged_data['earliest_start']
            
            # Add to actual time for this type
            actual_times[schedule_type] += attendance_duration
            actual_times['standard'] += attendance_duration
            
            # Add to daily attendance if the date is in our range
            if record_date in daily_attendance:
                daily_attendance[record_date] += attendance_duration
        
        # Format durations for Excel
        scheduled_times_formatted = {k: format_duration(v) for k, v in scheduled_times.items()}
        actual_times_formatted = {k: format_duration(v) for k, v in actual_times.items()}
        
        # Format daily attendance
        daily_attendance_formatted = {date_obj: format_duration(duration) if duration.total_seconds() > 0 else "-" 
                                     for date_obj, duration in daily_attendance.items()}
        
        # Calculate variance
        variance = scheduled_times['standard'] - actual_times['standard']
        variance_formatted = format_duration(abs(variance))
        if variance.total_seconds() < 0:
            variance_formatted = f"-{variance_formatted}"
        
        # Add row to Excel - main data
        data_row = [
            teacher_name,
            scheduled_times_formatted['standard'],
            scheduled_times_formatted['built_in'],
            scheduled_times_formatted['regular'],
            scheduled_times_formatted['overload'],
            actual_times_formatted['standard'],
            actual_times_formatted['built_in'],
            actual_times_formatted['regular'],
            actual_times_formatted['overload'],
            variance_formatted
        ]
        
        # Add daily attendance data
        for date_obj in all_dates:
            data_row.append(daily_attendance_formatted[date_obj])
        
        for col_num, value in enumerate(data_row, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="center")
            
            # Add borders to all cells
            cell.border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        row_num += 1
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Teacher_Hours_Summary_{start_date}_to_{end_date}.xlsx'
    
    # Save workbook to response
    workbook.save(response)
    
    return response


@login_required
def export_teacher_schedule_excel(request):
    """
    Export teacher schedule data to Excel based on the subjects they're assigned to
    """
    teacher_id = request.GET.get("teacher_id")
    semester_id = request.GET.get("semester_id")

    # Get teacher information
    if teacher_id:
        try:
            teacher_id_int = int(teacher_id)
            teacher = get_object_or_404(CustomUser, id=teacher_id_int)
        except (ValueError, TypeError):
            return HttpResponse("Invalid teacher ID", status=400)
    else:
        messages.error(request, "Teacher ID is required")
        return redirect('teacher_attendance_report')

    # Get current semester based on date range
    today_date = date.today()
    current_semester = Semester.objects.filter(
        start_date__lte=today_date,
        end_date__gte=today_date
    ).first()
    
    # Determine which semester to use for filtering
    selected_semester = None
    if semester_id:
        try:
            selected_semester = Semester.objects.get(id=int(semester_id))
        except (ValueError, TypeError, Semester.DoesNotExist):
            selected_semester = current_semester
    else:
        selected_semester = current_semester

    # Get all subjects assigned to this teacher
    subjects = Subject.objects.filter(assign_teacher=teacher)

    # Create Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Teacher Schedule"
    
    # Add header row
    headers = [
        "Subject", "Subject Code", "Room", "Days", "Start Time", "End Time", "Duration", "Schedule Type"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1E88E5", end_color="1E88E5", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal="center")
    
    # Add teacher information at the top
    teacher_name = f"{teacher.first_name} {teacher.last_name}"
    sheet.insert_rows(1, 2)
    sheet.merge_cells('A1:H1')
    title_cell = sheet.cell(row=1, column=1, value=f"Schedule for: {teacher_name}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Start data rows from row 4 (after title and headers)
    row_num = 4
    
    # Process schedule data for each subject
    for subject in subjects:
        schedules = Schedule.objects.filter(subject=subject)
        
        # Apply semester filter to schedules
        if selected_semester:
            schedules = schedules.filter(semester=selected_semester)
        
        for schedule in schedules:
            # Format days of week
            days_display = ", ".join(schedule.days_of_week)
            
            # Calculate duration
            start_time = schedule.schedule_start_time
            end_time = schedule.schedule_end_time
            
            # Calculate duration in hours and minutes
            start_datetime = datetime.combine(date.today(), start_time)
            end_datetime = datetime.combine(date.today(), end_time)
            duration = end_datetime - start_datetime
            hours = int(duration.total_seconds() // 3600)
            minutes = int((duration.total_seconds() % 3600) // 60)
            duration_str = f"{hours:02d}:{minutes:02d}"
            
            # Add row data
            sheet.cell(row=row_num, column=1, value=subject.subject_name)
            sheet.cell(row=row_num, column=2, value=subject.subject_code)
            sheet.cell(row=row_num, column=3, value=subject.room_number)
            sheet.cell(row=row_num, column=4, value=days_display)
            sheet.cell(row=row_num, column=5, value=format_time(start_time))
            sheet.cell(row=row_num, column=6, value=format_time(end_time))
            sheet.cell(row=row_num, column=7, value=duration_str)
            sheet.cell(row=row_num, column=8, value=schedule.schedule_type if schedule.schedule_type else "Regular")
            
            # Apply styling
            for col in range(1, 9):
                cell = sheet.cell(row=row_num, column=col)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Center align specific columns
                if col in [3, 4, 5, 6, 7, 8]:
                    cell.alignment = Alignment(horizontal="center")
            
            row_num += 1
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        sheet.column_dimensions[column_letter].width = 18
    
    # Create HTTP response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Teacher_Schedule_{teacher_name.replace(' ', '_')}.xlsx"
    workbook.save(response)
    return response
